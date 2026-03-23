import pandas as pd
import numpy as np
import streamlit as st
import io
import re
import logging
import zipfile
import openpyxl
from openpyxl.styles import Font, Alignment
from collections import defaultdict
from datetime import datetime
from itertools import combinations
import warnings
import traceback
import hashlib
from functools import lru_cache

# 配置日志和警告
warnings.filterwarnings('ignore')
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger('MultiAccountWashTrade')

# Streamlit 页面配置
st.set_page_config(
    page_title="🎈智能对刷检测系统🎈",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==================== 配置类 ====================
class Config:
    def __init__(self):
        self.min_amount = 10
        self.amount_similarity_threshold = 0.8
        self.min_continuous_periods = 3
        self.max_accounts_in_group = 5
        self.supported_file_types = ['.xlsx', '.xls', '.csv']
        
        # 列名映射配置
        self.column_mappings = {
            '会员账号': ['会员账号', '会员账户', '账号', '账户', '用户账号', '玩家账号', '用户ID', '玩家ID'],
            '彩种': ['彩种', '彩神', '彩票种类', '游戏类型', '彩票类型', '游戏彩种', '彩票名称'],
            '期号': ['期号', '期数', '期次', '期', '奖期', '期号信息', '期号编号'],
            '玩法': ['玩法', '玩法分类', '投注类型', '类型', '投注玩法', '玩法类型', '分类'],
            '内容': ['内容', '投注内容', '下注内容', '注单内容', '投注号码', '号码内容', '投注信息'],
            '金额': ['金额', '下注总额', '投注金额', '总额', '下注金额', '投注额', '金额数值']
        }
        
        # 活跃度阈值配置
        self.period_thresholds = {
            'low_activity': 10,
            'medium_activity_low': 11,
            'medium_activity_high': 50,
            'high_activity_low': 51,
            'high_activity_high': 100,
            'min_periods_low': 3,
            'min_periods_medium': 5,
            'min_periods_high': 8,
            'min_periods_very_high': 11
        }
        
        # 多账户匹配度阈值
        self.account_count_similarity_thresholds = {
            2: 0.7,
            3: 0.8,
            4: 0.85,
            5: 0.9
        }
        
        # 账户期数差异阈值
        self.account_period_diff_threshold = 101
        
        # 方向模式
        self.base_direction_patterns = {
            '小': ['两面-小', '和值-小', '小', 'small', 'xia', 'xiao', '和值小', '总和-小', '总和小', '和值_小', '总和_小'],
            '大': ['两面-大', '和值-大', '大', 'big', 'da', 'large', '和值大', '总和-大', '总和大', '和值_大', '总和_大'], 
            '单': ['两面-单', '和值-单', '单', 'odd', 'dan', '奇数', '和值单', '总和-单', '总和单', '和值_单', '总和_单'],
            '双': ['两面-双', '和值-双', '双', 'even', 'shuang', '偶数', '和值双', '总和-双', '总和双', '和值_双', '总和_双'],
            '龙': ['龙', 'long', 'dragon', '龍', '龍虎-龙', '龙虎-龙', '龍虎_龙'],
            '虎': ['虎', 'hu', 'tiger', '龍虎-虎', '龙虎-虎', '龍虎_虎'],
            '质': ['质', '质数', 'prime', 'zhi', '質', '質數', '素数'],
            '合': ['合', '合数', 'composite', 'he', '合數', '合成数'],
        }
        
        # 增强方向模式
        self.enhanced_direction_patterns = {
            '特小': ['特小', '极小', '最小', '特小单', '特小双', '特码-小', '特码小', '特码_小'],
            '特大': ['特大', '极大', '最大', '特单大', '特双大', '特码-大', '特码大', '特码_大'],
            '特单': ['特单', '特码-单', '特码单', '特码_单'],
            '特双': ['特双', '特码-双', '特码双', '特码_双'],
            
            # 添加总和龙虎格式
            '总和小': ['总和小', '和小', '总和-小', '和值小', '和值-小', '冠亚和小', '冠亚和-小', '总和、龙虎-总和小'],
            '总和大': ['总和大', '和大', '总和-大', '和值大', '和值-大', '冠亚和大', '冠亚和-大', '总和、龙虎-总和大'],
            '总和单': ['总和单', '和单', '总和-单', '和值单', '和值-单', '冠亚和单', '冠亚和-单', '总和、龙虎-总和单'],
            '总和双': ['总和双', '和双', '总和-双', '和值双', '和值-双', '冠亚和双', '冠亚和-双', '总和、龙虎-总和双'],
            
            '大单': ['大单', '单大', 'big-odd', '大-单', '单-大'],
            '大双': ['大双', '双大', 'big-even', '大-双', '双-大'],
            '小单': ['小单', '单小', 'small-odd', '小-单', '单-小'],
            '小双': ['小双', '双小', 'small-even', '小-双', '双-小'],
            
            '天肖': ['天肖', '天肖', '天', '天生肖', '天肖码'],
            '地肖': ['地肖', '地肖', '地', '地生肖', '地肖码'],
            '家肖': ['家肖', '家禽', '家肖', '家', '家禽肖', '家生肖'],
            '野肖': ['野肖', '野兽', '野肖', '野', '野兽肖', '野生肖'],
            '尾大': ['尾大', '尾大', '大尾', '尾数大', '尾數大'],
            '尾小': ['尾小', '尾小', '小尾', '尾数小', '尾數小'],
        
            '尾大': ['尾大', '尾大', '大尾', '尾数大', '尾數大', '特码两面-尾大'],
            '尾小': ['尾小', '尾小', '小尾', '尾数小', '尾數小', '特码两面-尾小'],
            '特大': ['特大', '极大', '最大', '特单大', '特双大', '特码-大', '特码大', '特码_大', '特码两面-特大'],
            '特小': ['特小', '极小', '最小', '特小单', '特小双', '特码-小', '特码小', '特码_小', '特码两面-特小'],
            '特单': ['特单', '特码-单', '特码单', '特码_单', '特码两面-特单'],
            '特双': ['特双', '特码-双', '特码双', '特码_双', '特码两面-特双'],
            
            '大': ['大', 'big', 'large', 'da', '特码两面-大'],
            '小': ['小', 'small', 'xiao', '特码两面-小'], 
            '单': ['单', 'odd', 'dan', '奇', '特码两面-单'],
            '双': ['双', 'even', 'shuang', '偶', '特码两面-双'],
            
            '正1特-大': ['正1特-大', '正一特-大', '正码特_正一特-大', '正1特大', '正一特大', '正码特-正一特-大'],
            '正1特-小': ['正1特-小', '正一特-小', '正码特_正一特-小', '正1特小', '正一特小', '正码特-正一特-小'],
            '正1特-单': ['正1特-单', '正一特-单', '正码特_正一特-单', '正1特单', '正一特单', '正码特-正一特-单'],
            '正1特-双': ['正1特-双', '正一特-双', '正码特_正一特-双', '正1特双', '正一特双', '正码特-正一特-双'],
            
            '正2特-大': ['正2特-大', '正二特-大', '正码特_正二特-大', '正2特大', '正二特大', '正码特-正二特-大'],
            '正2特-小': ['正2特-小', '正二特-小', '正码特_正二特-小', '正2特小', '正二特小', '正码特-正二特-小'],
            '正2特-单': ['正2特-单', '正二特-单', '正码特_正二特-单', '正2特单', '正二特单', '正码特-正二特-单'],
            '正2特-双': ['正2特-双', '正二特-双', '正码特_正二特-双', '正2特双', '正二特双', '正码特-正二特-双'],
            
            '正3特-大': ['正3特-大', '正三特-大', '正码特_正三特-大', '正3特大', '正三特大', '正码特-正三特-大'],
            '正3特-小': ['正3特-小', '正三特-小', '正码特_正三特-小', '正3特小', '正三特小', '正码特-正三特-小'],
            '正3特-单': ['正3特-单', '正三特-单', '正码特_正三特-单', '正3特单', '正三特单', '正码特-正三特-单'],
            '正3特-双': ['正3特-双', '正三特-双', '正码特_正三特-双', '正3特双', '正三特双', '正码特-正三特-双'],
            
            '正4特-大': ['正4特-大', '正四特-大', '正码特_正四特-大', '正4特大', '正四特大', '正码特-正四特-大'],
            '正4特-小': ['正4特-小', '正四特-小', '正码特_正四特-小', '正4特小', '正四特小', '正码特-正四特-小'],
            '正4特-单': ['正4特-单', '正四特-单', '正码特_正四特-单', '正4特单', '正四特单', '正码特-正四特-单'],
            '正4特-双': ['正4特-双', '正四特-双', '正码特_正四特-双', '正4特双', '正四特双', '正码特-正四特-双'],
            
            '正5特-大': ['正5特-大', '正五特-大', '正码特_正五特-大', '正5特大', '正五特大', '正码特-正五特-大'],
            '正5特-小': ['正5特-小', '正五特-小', '正码特_正五特-小', '正5特小', '正五特小', '正码特-正五特-小'],
            '正5特-单': ['正5特-单', '正五特-单', '正码特_正五特-单', '正5特单', '正五特单', '正码特-正五特-单'],
            '正5特-双': ['正5特-双', '正五特-双', '正码特_正五特-双', '正5特双', '正五特双', '正码特-正五特-双'],
            
            '正6特-大': ['正6特-大', '正六特-大', '正码特_正六特-大', '正6特大', '正六特大', '正码特-正六特-大'],
            '正6特-小': ['正6特-小', '正六特-小', '正码特_正六特-小', '正6特小', '正六特小', '正码特-正六特-小'],
            '正6特-单': ['正6特-单', '正六特-单', '正码特_正六特-单', '正6特单', '正六特单', '正码特-正六特-单'],
            '正6特-双': ['正6特-双', '正六特-双', '正码特_正六特-双', '正6特双', '正六特双', '正码特-正六特-双'],
            
            '正1-大': ['正1-大', '正码1-大', '正一-大', '正码_正一-大', '正码1大', '正一大'],
            '正1-小': ['正1-小', '正码1-小', '正一-小', '正码_正一-小', '正码1小', '正一小'],
            '正1-单': ['正1-单', '正码1-单', '正一-单', '正码_正一-单', '正码1单', '正一单'],
            '正1-双': ['正1-双', '正码1-双', '正一-双', '正码_正一-双', '正码1双', '正一双'],
            
            '正2-大': ['正2-大', '正码2-大', '正二-大', '正码_正二-大', '正码2大', '正二大'],
            '正2-小': ['正2-小', '正码2-小', '正二-小', '正码_正二-小', '正码2小', '正二小'],
            '正2-单': ['正2-单', '正码2-单', '正二-单', '正码_正二-单', '正码2单', '正二单'],
            '正2-双': ['正2-双', '正码2-双', '正二-双', '正码_正二-双', '正码2双', '正二双'],
            
            '正3-大': ['正3-大', '正码3-大', '正三-大', '正码_正三-大', '正码3大', '正三大'],
            '正3-小': ['正3-小', '正码3-小', '正三-小', '正码_正三-小', '正码3小', '正三小'],
            '正3-单': ['正3-单', '正码3-单', '正三-单', '正码_正三-单', '正码3单', '正三单'],
            '正3-双': ['正3-双', '正码3-双', '正三-双', '正码_正三-双', '正码3双', '正三双'],
            
            '正4-大': ['正4-大', '正码4-大', '正四-大', '正码_正四-大', '正码4大', '正四大'],
            '正4-小': ['正4-小', '正码4-小', '正四-小', '正码_正四-小', '正码4小', '正四小'],
            '正4-单': ['正4-单', '正码4-单', '正四-单', '正码_正四-单', '正码4单', '正四单'],
            '正4-双': ['正4-双', '正码4-双', '正四-双', '正码_正四-双', '正码4双', '正四双'],
            
            '正5-大': ['正5-大', '正码5-大', '正五-大', '正码_正五-大', '正码5大', '正五大'],
            '正5-小': ['正5-小', '正码5-小', '正五-小', '正码_正五-小', '正码5小', '正五小'],
            '正5-单': ['正5-单', '正码5-单', '正五-单', '正码_正五-单', '正码5单', '正五单'],
            '正5-双': ['正5-双', '正码5-双', '正五-双', '正码_正五-双', '正码5双', '正五双'],
            
            '正6-大': ['正6-大', '正码6-大', '正六-大', '正码_正六-大', '正码6大', '正六大'],
            '正6-小': ['正6-小', '正码6-小', '正六-小', '正码_正六-小', '正码6小', '正六小'],
            '正6-单': ['正6-单', '正码6-单', '正六-单', '正码_正六-单', '正码6单', '正六单'],
            '正6-双': ['正6-双', '正码6-双', '正六-双', '正码_正六-双', '正码6双', '正六双'],
        }
        
        # 合并方向模式
        self.direction_patterns = {**self.base_direction_patterns, **self.enhanced_direction_patterns}
        
        # 对立组配置
        self.opposite_groups = [
            {'大', '小'}, {'单', '双'}, {'龙', '虎'}, {'质', '合'},
            {'特大', '特小'}, {'特单', '特双'}, 
            {'总和大', '总和小'}, {'总和单', '总和双'},
            {'大单', '小双'}, {'大双', '小单'},
            {'天肖', '地肖'}, {'家肖', '野肖'}, {'尾大', '尾小'},
            
            {'正1特-大', '正1特-小'}, {'正1特-单', '正1特-双'},
            {'正2特-大', '正2特-小'}, {'正2特-单', '正2特-双'},
            {'正3特-大', '正3特-小'}, {'正3特-单', '正3特-双'},
            {'正4特-大', '正4特-小'}, {'正4特-单', '正4特-双'},
            {'正5特-大', '正5特-小'}, {'正5特-单', '正5特-双'},
            {'正6特-大', '正6特-小'}, {'正6特-单', '正6特-双'},
            
            {'正1-大', '正1-小'}, {'正1-单', '正1-双'},
            {'正2-大', '正2-小'}, {'正2-单', '正2-双'},
            {'正3-大', '正3-小'}, {'正3-单', '正3-双'},
            {'正4-大', '正4-小'}, {'正4-单', '正4-双'},
            {'正5-大', '正5-小'}, {'正5-单', '正5-双'},
            {'正6-大', '正6-小'}, {'正6-单', '正6-双'},
            
            {'尾大', '尾小'},
            {'特大', '特小'},
            {'特单', '特双'},
            {'特码两面-尾大', '特码两面-尾小'},
            {'特码两面-特大', '特码两面-特小'},
            {'特码两面-特单', '特码两面-特双'},
        ]
        
        # 位置关键词映射
        self.position_keywords = {
            'PK10': {
                '冠军': ['冠军', '第1名', '第一名', '前一', '冠 军', '冠　军'],
                '亚军': ['亚军', '第2名', '第二名', '亚 军', '亚　军'],
                '季军': ['季军', '第3名', '第三名', '季 军', '季　军'],
                '第四名': ['第四名', '第4名'],
                '第五名': ['第五名', '第5名'],
                '第六名': ['第六名', '第6名'],
                '第七名': ['第七名', '第7名'],
                '第八名': ['第八名', '第8名'],
                '第九名': ['第九名', '第9名'],
                '第十名': ['第十名', '第10名']
            },
            '3D': {
                '百位': ['百位', '定位_百位', '百位定位'],
                '十位': ['十位', '定位_十位', '十位定位'],
                '个位': ['个位', '定位_个位', '个位定位']
            },
            'SSC': {
                '第1球': ['第1球', '万位', '第一位', '定位_万位', '万位定位'],
                '第2球': ['第2球', '千位', '第二位', '定位_千位', '千位定位'],
                '第3球': ['第3球', '百位', '第三位', '定位_百位', '百位定位'],
                '第4球': ['第4球', '十位', '第四位', '定位_十位', '十位定位'],
                '第5球': ['第5球', '个位', '第五位', '定位_个位', '个位定位']
            },
            'LHC': {
                '特码': ['特码', '特肖', '正码特', '特码A', '特码B'],
                '正1特': ['正1特', '正一特', '正码特_正一特', '正码特-正一特'],
                '正2特': ['正2特', '正二特', '正码特_正二特', '正码特-正二特'],
                '正3特': ['正3特', '正三特', '正码特_正三特', '正码特-正三特'],
                '正4特': ['正4特', '正四特', '正码特_正四特', '正码特-正四特'],
                '正5特': ['正5特', '正五特', '正码特_正五特', '正码特-正五特'],
                '正6特': ['正6特', '正六特', '正码特_正六特', '正码特-正六特'],
                '正1': ['正1', '正一', '正码1', '正码_正一'],
                '正2': ['正2', '正二', '正码2', '正码_正二'],
                '正3': ['正3', '正三', '正码3', '正码_正三'],
                '正4': ['正4', '正四', '正码4', '正码_正四'],
                '正5': ['正5', '正五', '正码5', '正码_正五'],
                '正6': ['正6', '正六', '正码6', '正码_正六'],
                '平特': ['平特', '平特肖', '平码'],
                '连肖': ['连肖', '二连肖', '三连肖', '四连肖'],
                '连尾': ['连尾', '二连尾', '三连尾', '四连尾'],
                '色波': ['色波', '红波', '蓝波', '绿波'],
                '五行': ['五行', '金', '木', '水', '火', '土']
            }
        }

        # 金额阈值配置
        self.amount_threshold = {
            'max_amount_ratio': 10,
            'enable_threshold_filter': True
        }

# ==================== 数据处理器类 ====================
class DataProcessor:
    def __init__(self):
        self.required_columns = ['会员账号', '彩种', '期号', '玩法', '内容', '金额']
        self.column_mapping = {
            '会员账号': ['会员账号', '会员账户', '账号', '账户', '用户账号', '玩家账号', '用户ID', '玩家ID', '用户名称', '玩家名称'],
            '彩种': ['彩种', '彩神', '彩票种类', '游戏类型', '彩票类型', '游戏彩种', '彩票名称', '彩系', '游戏名称'],
            '期号': ['期号', '期数', '期次', '期', '奖期', '期号信息', '期号编号', '开奖期号', '奖期号'],
            '玩法': ['玩法', '玩法分类', '投注类型', '类型', '投注玩法', '玩法类型', '分类', '玩法名称', '投注方式'],
            '内容': ['内容', '投注内容', '下注内容', '注单内容', '投注号码', '号码内容', '投注信息', '号码', '选号'],
            '金额': ['金额', '下注总额', '投注金额', '总额', '下注金额', '投注额', '金额数值', '单注金额', '投注额', '钱', '元']
        }
        
        self.similarity_threshold = 0.7
    
    def smart_column_identification(self, df_columns):
        """智能列识别"""
        identified_columns = {}
        actual_columns = [str(col).strip() for col in df_columns]
        
        for standard_col, possible_names in self.column_mapping.items():
            found = False
            for actual_col in actual_columns:
                actual_col_lower = actual_col.lower().replace(' ', '').replace('_', '').replace('-', '')
                
                for possible_name in possible_names:
                    possible_name_lower = possible_name.lower().replace(' ', '').replace('_', '').replace('-', '')
                    
                    set1 = set(possible_name_lower)
                    set2 = set(actual_col_lower)
                    intersection = set1 & set2
                    
                    similarity_score = len(intersection) / len(set1) if set1 else 0
                    
                    if (possible_name_lower in actual_col_lower or 
                        actual_col_lower in possible_name_lower or
                        similarity_score >= self.similarity_threshold):
                        
                        identified_columns[actual_col] = standard_col
                        found = True
                        break
                
                if found:
                    break
        
        return identified_columns
    
    def find_data_start(self, df):
        """智能找到数据起始位置"""
        for row_idx in range(min(20, len(df))):
            for col_idx in range(min(10, len(df.columns))):
                cell_value = str(df.iloc[row_idx, col_idx])
                if pd.notna(cell_value) and any(keyword in cell_value for keyword in ['会员', '账号', '期号', '彩种', '玩法', '内容', '订单', '用户']):
                    return row_idx, col_idx
        return 0, 0
    
    def validate_data_quality(self, df):
        """数据质量验证"""
        logger.info("正在进行数据质量验证...")
        issues = []
        
        # 检查必要列
        missing_cols = [col for col in self.required_columns if col not in df.columns]
        if missing_cols:
            issues.append(f"缺少必要列: {missing_cols}")
        
        # 检查空值
        for col in self.required_columns:
            if col in df.columns:
                null_count = df[col].isnull().sum()
                if null_count > 0:
                    issues.append(f"列 '{col}' 有 {null_count} 个空值")

        # 检查重复数据
        duplicate_count = df.duplicated().sum()
        if duplicate_count > 0:
            issues.append(f"发现 {duplicate_count} 条重复记录")

        return issues
    
    def clean_data(self, uploaded_file):
        """数据清洗主函数"""
        try:
            df_temp = pd.read_excel(uploaded_file, header=None, nrows=50)
            
            start_row, start_col = self.find_data_start(df_temp)
            
            df_clean = pd.read_excel(
                uploaded_file, 
                header=start_row,
                skiprows=range(start_row + 1) if start_row > 0 else None,
                dtype=str,
                na_filter=False,
                keep_default_na=False
            )
            
            if start_col > 0:
                df_clean = df_clean.iloc[:, start_col:]
            
            column_mapping = self.smart_column_identification(df_clean.columns)
            if column_mapping:
                df_clean = df_clean.rename(columns=column_mapping)
            
            missing_columns = [col for col in self.required_columns if col not in df_clean.columns]
            if missing_columns and len(df_clean.columns) >= 4:
                manual_mapping = {}
                col_names = ['会员账号', '彩种', '期号', '内容', '玩法', '金额']
                for i, col_name in enumerate(col_names):
                    if i < len(df_clean.columns):
                        manual_mapping[df_clean.columns[i]] = col_name
                
                df_clean = df_clean.rename(columns=manual_mapping)
            
            initial_count = len(df_clean)
            df_clean = df_clean.dropna(subset=[col for col in self.required_columns if col in df_clean.columns])
            df_clean = df_clean.dropna(axis=1, how='all')
            
            for col in self.required_columns:
                if col in df_clean.columns:
                    if col == '会员账号':
                        df_clean[col] = df_clean[col].apply(
                            lambda x: str(x) if pd.notna(x) else ''
                        )
                    else:
                        df_clean[col] = df_clean[col].astype(str).str.strip()
            
            if '期号' in df_clean.columns:
                df_clean['期号'] = df_clean['期号'].str.replace(r'\.0$', '', regex=True)
            
            if '金额' in df_clean.columns:
                df_clean['金额'] = df_clean['金额'].apply(self.preprocess_amount_column)
            
            if '内容' in df_clean.columns:
                df_clean['内容'] = df_clean['内容'].apply(self.preprocess_content_column)
            
            self.validate_data_quality(df_clean)
            
            return df_clean
                
        except Exception as e:
            st.error(f"❌ 数据清洗失败: {str(e)}")
            logger.error(f"数据清洗失败: {str(e)}")
            return None
    
    def preprocess_amount_column(self, amount_text):
        """预处理金额列格式"""
        if pd.isna(amount_text):
            return amount_text
        
        text = str(amount_text).strip()
        
        if '投注：' in text and '抵用：' in text:
            try:
                bet_part = text.split('投注：')[1].split('抵用：')[0].strip()
                return f"投注：{bet_part}"
            except:
                return text
        
        return text
    
    def preprocess_content_column(self, content_text):
        """预处理内容列格式"""
        if pd.isna(content_text):
            return content_text
        
        text = str(content_text).strip()
        
        if '特码两面-' in text:
            text = text.replace('特码两面 - ', '特码两面-')
            text = text.replace('特码两面- ', '特码两面-')
            return text
        
        return text

# ==================== 彩种识别器 ====================
LOTTERY_CONFIGS = {
    'PK10': {
        'lotteries': [
            '分分PK拾', '三分PK拾', '五分PK拾', '新幸运飞艇', '澳洲幸运10',
            '一分PK10', '宾果PK10', '极速飞艇', '澳洲飞艇', '幸运赛车',
            '分分赛车', '北京PK10', '旧北京PK10', '极速赛车', '幸运赛車', 
            '北京赛车', '极速PK10', '幸运PK10', '赛车', '赛車'
        ],
        'min_number': 1,
        'max_number': 10,
        'gyh_min': 3,
        'gyh_max': 19,
        'position_names': ['冠军', '亚军', '第三名', '第四名', '第五名', 
                          '第六名', '第七名', '第八名', '第九名', '第十名']
    },
    'K3': {
        'lotteries': [
            '分分快三', '三分快3', '五分快3', '澳洲快三', '宾果快三',
            '1分快三', '3分快三', '5分快三', '10分快三', '加州快三',
            '幸运快三', '大发快三', '快三', '快3', 'k3', 'k三', 
            '澳门快三', '香港快三', '江苏快三'
        ],
        'min_number': 1,
        'max_number': 6,
        'hezhi_min': 3,
        'hezhi_max': 18
    },
    'LHC': {
        'lotteries': [
            '新澳门六合彩', '澳门六合彩', '香港六合彩', '一分六合彩',
            '五分六合彩', '三分六合彩', '香港⑥合彩', '分分六合彩',
            '快乐6合彩', '港⑥合彩', '台湾大乐透', '六合', 'lhc', '六合彩',
            '⑥合', '6合', '大发六合彩'
        ],
        'min_number': 1,
        'max_number': 49
    },
    'SSC': {
        'lotteries': [
            '分分时时彩', '三分时时彩', '五分时时彩', '宾果时时彩',
            '1分时时彩', '3分时时彩', '5分时时彩', '旧重庆时时彩',
            '幸运时时彩', '腾讯分分彩', '新疆时时彩', '天津时时彩',
            '重庆时时彩', '上海时时彩', '广东时时彩', '分分彩', '时时彩', '時時彩'
        ],
        'min_number': 0,
        'max_number': 9
    },
    '3D': {
        'lotteries': [
            '排列三', '排列3', '幸运排列3', '一分排列3', '二分排列3', '三分排列3', 
            '五分排列3', '十分排列3', '大发排列3', '好运排列3', '福彩3D', '极速3D',
            '极速排列3', '幸运3D', '一分3D', '二分3D', '三分3D', '五分3D', 
            '十分3D', '大发3D', '好运3D'
        ],
        'min_number': 0,
        'max_number': 9,
        'position_names': ['百位', '十位', '个位']
    }
}

class LotteryIdentifier:
    def __init__(self):
        self.lottery_configs = LOTTERY_CONFIGS
        self.general_keywords = {
            'PK10': ['pk10', 'pk拾', '飞艇', '赛车', '赛車', '幸运10', '北京赛车', '极速赛车'],
            'K3': ['快三', '快3', 'k3', 'k三', '骰宝', '三军'],
            'LHC': ['六合', 'lhc', '六合彩', '⑥合', '6合', '特码', '平特', '连肖'],
            'SSC': ['时时彩', 'ssc', '分分彩', '時時彩', '重庆时时彩', '腾讯分分彩'],
            '3D': ['排列三', '排列3', '福彩3d', '3d', '极速3d', '排列', 'p3', 'p三']
        }
        
        self.lottery_aliases = {
            '分分PK拾': 'PK10', '三分PK拾': 'PK10', '五分PK拾': 'PK10',
            '新幸运飞艇': 'PK10', '澳洲幸运10': 'PK10', '一分PK10': 'PK10',
            '宾果PK10': 'PK10', '极速飞艇': 'PK10', '澳洲飞艇': 'PK10',
            '幸运赛车': 'PK10', '分分赛车': 'PK10', '北京PK10': 'PK10',
            '旧北京PK10': 'PK10', '极速赛车': 'PK10', '幸运赛車': 'PK10',
            '北京赛车': 'PK10', '极速PK10': 'PK10', '幸运PK10': 'PK10',
            '分分快三': 'K3', '三分快3': 'K3', '五分快3': 'K3', '澳洲快三': 'K3',
            '宾果快三': 'K3', '1分快三': 'K3', '3分快三': 'K3', '5分快三': 'K3',
            '10分快三': 'K3', '加州快三': 'K3', '幸运快三': 'K3', '大发快三': 'K3',
            '澳门快三': 'K3', '香港快三': 'K3', '江苏快三': 'K3',
            '新澳门六合彩': 'LHC', '澳门六合彩': 'LHC', '香港六合彩': 'LHC',
            '一分六合彩': 'LHC', '五分六合彩': 'LHC', '三分六合彩': 'LHC',
            '香港⑥合彩': 'LHC', '分分六合彩': 'LHC', '快乐6合彩': 'LHC',
            '港⑥合彩': 'LHC', '台湾大乐透': 'LHC', '大发六合彩': 'LHC',
            '分分时时彩': 'SSC', '三分时时彩': 'SSC', '五分时时彩': 'SSC',
            '宾果时时彩': 'SSC', '1分时时彩': 'SSC', '3分时时彩': 'SSC',
            '5分时时彩': 'SSC', '旧重庆时时彩': 'SSC', '幸运时时彩': 'SSC',
            '腾讯分分彩': 'SSC', '新疆时时彩': 'SSC', '天津时时彩': 'SSC',
            '重庆时时彩': 'SSC', '上海时时彩': 'SSC', '广东时时彩': 'SSC',
            '排列三': '3D', '排列3': '3D', '幸运排列3': '3D', '一分排列3': '3D',
            '二分排列3': '3D', '三分排列3': '3D', '五分排列3': '3D', '十分排列3': '3D',
            '大发排列3': '3D', '好运排列3': '3D', '福彩3D': '3D', '极速3D': '3D',
            '极速排列3': '3D', '幸运3D': '3D', '一分3D': '3D', '二分3D': '3D',
            '三分3D': '3D', '五分3D': '3D', '十分3D': '3D', '大发3D': '3D', '好运3D': '3D'
        }

    def identify_lottery_type(self, lottery_name):
        """彩种类型识别"""
        lottery_str = str(lottery_name).strip()
        
        if lottery_str in self.lottery_aliases:
            return self.lottery_aliases[lottery_str]
        
        for lottery_type, config in self.lottery_configs.items():
            for lottery in config['lotteries']:
                if lottery in lottery_str:
                    return lottery_type
        
        lottery_lower = lottery_str.lower()
        
        for lottery_type, keywords in self.general_keywords.items():
            for keyword in keywords:
                if keyword.lower() in lottery_lower:
                    return lottery_type
        
        return lottery_str

# ==================== 玩法分类器 ====================
class PlayCategoryNormalizer:
    def __init__(self):
        self.category_mapping = self._create_category_mapping()
    
    def _create_category_mapping(self):
        """创建玩法分类映射"""
        mapping = {
            '和值': '和值', '和值_大小单双': '和值', '两面': '两面',
            '二不同号': '二不同号', '三不同号': '三不同号', '独胆': '独胆',
            '点数': '和值', '三军': '独胆', '三軍': '独胆',
            
            '特码': '特码', '正1特': '正1特', '正码特_正一特': '正1特',
            '正2特': '正2特', '正码特_正二特': '正2特', '正3特': '正3特',
            '正码特_正三特': '正3特', '正4特': '正4特', '正码特_正四特': '正4特',
            '正5特': '正5特', '正码特_正五特': '正5特', '正6特': '正6特',
            '正码特_正六特': '正6特', '正码': '正码', '正特': '正特',
            '尾数': '尾数', '特肖': '特肖', '平特': '平特', '一肖': '一肖',
            '连肖': '连肖', '连尾': '连尾', '龙虎': '龙虎', '五行': '五行',
            '色波': '色波', '半波': '半波', '天肖': '天肖', '地肖': '地肖',
            '家肖': '家肖', '野肖': '野肖',
    
            '正1特': '正1特', '正码特_正一特': '正1特', '正码特-正一特': '正1特',
            '正2特': '正2特', '正码特_正二特': '正2特', '正码特-正二特': '正2特',
            '正3特': '正3特', '正码特_正三特': '正3特', '正码特-正三特': '正3特',
            '正4特': '正4特', '正码特_正四特': '正4特', '正码特-正四特': '正4特',
            '正5特': '正5特', '正码特_正五特': '正5特', '正码特-正五特': '正5特',
            '正6特': '正6特', '正码特_正六特': '正6特', '正码特-正六特': '正6特',
            
            '正1': '正1', '正码1': '正1', '正码_正一': '正1',
            '正2': '正2', '正码2': '正2', '正码_正二': '正2',
            '正3': '正3', '正码3': '正3', '正码_正三': '正3',
            '正4': '正4', '正码4': '正4', '正码_正四': '正4',
            '正5': '正5', '正码5': '正5', '正码_正五': '正5',
            '正6': '正6', '正码6': '正6', '正码_正六': '正6',
            
            '两面': '两面', '大小单双': '两面', '百位': '百位', '十位': '十位', 
            '个位': '个位', '百十': '百十', '百个': '百个', '十个': '十个',
            '百十个': '百十个', '定位胆': '定位胆', '定位胆_百位': '定位胆_百位',
            '定位胆_十位': '定位胆_十位', '定位胆_个位': '定位胆_个位',
            
            '斗牛': '斗牛', '1-5球': '1-5球', '第1球': '第1球', '第2球': '第2球',
            '第3球': '第3球', '第4球': '第4球', '第5球': '第5球', '总和': '总和',
            '正码': '正码', '定位胆': '定位胆',
            
            '前一': '冠军', '定位胆': '定位胆', '1-5名': '1-5名', '6-10名': '6-10名',
            '冠军': '冠军', '亚军': '亚军', '季军': '第三名', '第3名': '第三名',
            '第三名': '第三名', '第4名': '第四名', '第四名': '第四名',
            '第5名': '第五名', '第五名': '第五名', '第6名': '第六名', '第六名': '第六名',
            '第7名': '第七名', '第七名': '第七名', '第8名': '第八名', '第八名': '第八名',
            '第9名': '第九名', '第九名': '第九名', '第10名': '第十名', '第十名': '第十名',
            '双面': '两面', '冠亚和': '冠亚和',

            # PK10定位胆玩法映射
            '定位胆_第1~5名': '1-5名',
            '定位胆_第6~10名': '6-10名',
            '定位胆_第1~5名定位胆': '1-5名',
            '定位胆_第6~10名定位胆': '6-10名',
            '定位胆1-5名': '1-5名',
            '定位胆6-10名': '6-10名',
            '第1~5名定位胆': '1-5名',
            '第6~10名定位胆': '6-10名',
            '定位胆_冠军': '冠军',
            '定位胆_亚军': '亚军',
            '定位胆_季军': '第三名',
            
            '1-5名': '1-5名',
            '6-10名': '6-10名', 
            '1-5名定位胆': '1-5名',
            '6-10名定位胆': '6-10名',
            '前一': '冠军',
            '前二': '亚军', 
            '前三': '第三名',
            '前四': '第四名',
            '前五': '第五名',
            '定位胆': '定位胆'
        }
        return mapping
    
    def normalize_category(self, category):
        """统一玩法分类名称"""
        category_str = str(category).strip()
        
        if category_str in self.category_mapping:
            return self.category_mapping[category_str]
        
        for key, value in self.category_mapping.items():
            if key in category_str:
                return value
        
        category_lower = category_str.lower()
        
        pk10_position_mapping = {
            '冠军': ['冠军', '第一名', '第1名', '1st', '前一'],
            '亚军': ['亚军', '第二名', '第2名', '2nd', '前二'], 
            '第三名': ['第三名', '第3名', '季军', '3rd', '前三'],
            '第四名': ['第四名', '第4名', '4th', '前四'],
            '第五名': ['第五名', '第5名', '5th', '前五'],
            '第六名': ['第六名', '第6名', '6th'],
            '第七名': ['第七名', '第7名', '7th'],
            '第八名': ['第八名', '第8名', '8th'],
            '第九名': ['第九名', '第9名', '9th'],
            '第十名': ['第十名', '第10名', '10th']
        }
        
        for position, keywords in pk10_position_mapping.items():
            for keyword in keywords:
                if keyword in category_lower:
                    return position
        
        if any(word in category_lower for word in ['百位']):
            return '百位'
        if any(word in category_lower for word in ['十位']):
            return '十位'
        if any(word in category_lower for word in ['个位']):
            return '个位'
        
        if any(word in category_lower for word in ['第1球', '万位']):
            return '第1球'
        if any(word in category_lower for word in ['第2球', '千位']):
            return '第2球'
        if any(word in category_lower for word in ['第3球', '百位']):
            return '第3球'
        if any(word in category_lower for word in ['第4球', '十位']):
            return '第4球'
        if any(word in category_lower for word in ['第5球', '个位']):
            return '第5球'
        
        if any(word in category_lower for word in ['天肖']):
            return '天肖'
        if any(word in category_lower for word in ['地肖']):
            return '地肖'
        if any(word in category_lower for word in ['家肖', '家禽']):
            return '家肖'
        if any(word in category_lower for word in ['野肖', '野兽']):
            return '野肖'
    
        if any(word in category_lower for word in ['正1特', '正一特']):
            return '正1特'
        if any(word in category_lower for word in ['正2特', '正二特']):
            return '正2特'
        if any(word in category_lower for word in ['正3特', '正三特']):
            return '正3特'
        if any(word in category_lower for word in ['正4特', '正四特']):
            return '正4特'
        if any(word in category_lower for word in ['正5特', '正五特']):
            return '正5特'
        if any(word in category_lower for word in ['正6特', '正六特']):
            return '正6特'
        
        if any(word in category_lower for word in ['正1', '正一']):
            return '正1'
        if any(word in category_lower for word in ['正2', '正二']):
            return '正2'
        if any(word in category_lower for word in ['正3', '正三']):
            return '正3'
        if any(word in category_lower for word in ['正4', '正四']):
            return '正4'
        if any(word in category_lower for word in ['正5', '正五']):
            return '正5'
        if any(word in category_lower for word in ['正6', '正六']):
            return '正6'
        
        return category_str

# ==================== 内容解析器 ====================
class ContentParser:
    """内容解析器 - 全面增强版，支持数字、方向、复杂格式"""

    @staticmethod
    def extract_basic_directions(content, config):
        """提取基础方向"""
        content_str = str(content).strip()
        directions = []
        
        if not content_str:
            return directions
        
        content_lower = content_str.lower()
        
        for direction, patterns in config.direction_patterns.items():
            for pattern in patterns:
                pattern_lower = pattern.lower()
                if (pattern_lower == content_lower or 
                    pattern_lower in content_lower or 
                    content_lower in pattern_lower):
                    directions.append(direction)
                    break
        
        return directions

    @staticmethod
    def extract_sum_and_dragon_tiger(content, config):
        """专门解析总和龙虎格式：总和、龙虎-总和双"""
        try:
            if pd.isna(content):
                return []
            
            content_str = str(content).strip()
            directions = []
            
            # 格式：总和、龙虎-总和双
            if '总和、龙虎-' in content_str:
                direction_part = content_str.split('总和、龙虎-')[-1].strip()
                
                # 检查是否是总和方向
                sum_keywords = {
                    '总和双': ['总和双', '和双', '总和-双', '和值双', '总和、龙虎-总和双'],
                    '总和单': ['总和单', '和单', '总和-单', '和值单', '总和、龙虎-总和单'],
                    '总和小': ['总和小', '和小', '总和-小', '和值小', '总和、龙虎-总和小'],
                    '总和大': ['总和大', '和大', '总和-大', '和值大', '总和、龙虎-总和大']
                }
                
                for direction, patterns in sum_keywords.items():
                    for pattern in patterns:
                        if pattern in direction_part or direction_part == pattern:
                            directions.append(direction)
                            break
            
            return directions
                
        except Exception as e:
            logger.warning(f"总和龙虎解析失败: {content}, 错误: {e}")
            return []

    @staticmethod
    def enhanced_extract_directions(content, config):
        """全面增强版方向提取"""
        try:
            if pd.isna(content):
                return []
            
            content_str = str(content).strip()
            
            # 1. 首先处理特殊格式：总和、龙虎-总和双
            if '总和、龙虎-' in content_str:
                sum_directions = ContentParser.extract_sum_and_dragon_tiger(content_str, config)
                if sum_directions:
                    return sum_directions
            
            # 2. 处理特码两面格式
            if '特码两面-' in content_str:
                direction_part = content_str.split('特码两面-')[-1].strip()
                for direction, patterns in config.direction_patterns.items():
                    for pattern in patterns:
                        if direction_part == pattern or direction_part in pattern:
                            return [direction]
            
            # 3. LHC特殊模式处理
            lhc_special_patterns = {
                '特码两面-尾大': '尾大',
                '特码两面-尾小': '尾小', 
                '特码两面-特大': '特大',
                '特码两面-特小': '特小',
                '特码两面-特单': '特单',
                '特码两面-特双': '特双',
                '特码两面-大': '大',
                '特码两面-小': '小',
                '特码两面-单': '单',
                '特码两面-双': '双'
            }
            
            for pattern, direction in lhc_special_patterns.items():
                if pattern in content_str:
                    return [direction]
            
            # 4. 预处理内容
            content_clean = ContentParser.preprocess_content(content_str)
            
            # 5. 多层级方向提取
            directions = set()
            
            # 5.1 精确匹配
            for direction, patterns in config.direction_patterns.items():
                for pattern in patterns:
                    if pattern == content_clean:
                        directions.add(direction)
                        break
            
            # 5.2 部分匹配
            if not directions:
                for direction, patterns in config.direction_patterns.items():
                    for pattern in patterns:
                        if pattern in content_clean:
                            directions.add(direction)
            
            # 5.3 智能LHC位置提取
            if not directions:
                directions = ContentParser.smart_lhc_position_extraction(content_clean, config)
            
            # 6. 提取数字（如果没有找到方向）
            if not directions:
                numbers = ContentParser.extract_all_numbers(content_str)
                if numbers:
                    if len(numbers) > 1:
                        unique_numbers = sorted(set(numbers))
                        return [f"多数字-{','.join(unique_numbers)}"]
                    else:
                        return [f"数字-{numbers[0]}"]
            
            return list(directions)
                
        except Exception as e:
            logger.warning(f"方向提取失败: {content}, 错误: {e}")
            return []

    @staticmethod
    def extract_all_numbers(content):
        """提取所有数字"""
        try:
            if pd.isna(content):
                return []
            
            content_str = str(content).strip()
            numbers = re.findall(r'\b\d{1,2}\b', content_str)
            
            valid_numbers = []
            for num in numbers:
                if num.isdigit():
                    num_int = int(num)
                    if 1 <= num_int <= 49:
                        valid_numbers.append(num)
            
            return list(set(valid_numbers))
        except:
            return []

    @staticmethod
    def parse_complex_content(content, play_category):
        """解析复杂内容格式"""
        try:
            if pd.isna(content):
                return {'type': 'unknown', 'value': ''}
            
            content_str = str(content).strip()
            
            # 处理位置-方向-数字格式：第三名-06,第四名-06,...
            if ',' in content_str and any(pos in content_str for pos in ['冠军', '亚军', '第']):
                items = content_str.split(',')
                positions = []
                values = []
                
                for item in items:
                    item_clean = item.strip()
                    if '-' in item_clean:
                        parts = item_clean.split('-')
                        if len(parts) >= 2:
                            position = parts[0].strip()
                            value = parts[1].strip()
                            
                            positions.append(position)
                            values.append(value)
                
                if len(set(values)) == 1:
                    return {
                        'type': 'multiple_positions', 
                        'value': values[0], 
                        'positions': positions,
                        'values': values
                    }
                else:
                    return {
                        'type': 'mixed_positions',
                        'value': '混合',
                        'positions': positions,
                        'values': values
                    }
            
            # 处理单个位置-值格式
            if '-' in content_str:
                parts = content_str.split('-')
                if len(parts) >= 2:
                    position = parts[0].strip()
                    value = parts[1].strip()
                    
                    return {
                        'type': 'single_position',
                        'position': position,
                        'value': value
                    }
            
            # 提取数字
            numbers = ContentParser.extract_all_numbers(content_str)
            if numbers:
                return {'type': 'number', 'value': numbers[0], 'values': numbers}
            
            return {'type': 'raw', 'value': content_str}
            
        except Exception as e:
            logger.warning(f"复杂内容解析失败: {content}, 错误: {e}")
            return {'type': 'error', 'value': content_str}

    @staticmethod
    def preprocess_content(content):
        """内容预处理"""
        content_str = str(content).strip()
        
        # 替换中文标点为英文标点
        content_str = content_str.replace('，', ',').replace('；', ';').replace('：', ':')
        
        # 压缩多余空格
        content_str = re.sub(r'\s+', ' ', content_str).strip()
        
        # 移除括号
        content_str = re.sub(r'[\(\)（）【】]', '', content_str)
        
        return content_str

    @staticmethod
    def multi_level_direction_extraction(content, config):
        """多层级方向提取"""
        directions = set()
        
        for direction, patterns in config.direction_patterns.items():
            for pattern in patterns:
                if pattern == content:
                    directions.add(direction)
                    break
        
        if not directions:
            for direction, patterns in config.direction_patterns.items():
                for pattern in patterns:
                    if pattern in content:
                        directions.add(direction)
        
        if not directions:
            directions = ContentParser.smart_lhc_position_extraction(content, config)
        
        return list(directions)

    @staticmethod
    def smart_lhc_position_extraction(content, config):
        """智能六合彩位置提取"""
        directions = set()
        content_lower = content.lower()
        
        lhc_position_map = {
            '正1特': ['正1特', '正一特', '正码特_正一特'],
            '正2特': ['正2特', '正二特', '正码特_正二特'], 
            '正3特': ['正3特', '正三特', '正码特_正三特'],
            '正4特': ['正4特', '正四特', '正码特_正四特'],
            '正5特': ['正5特', '正五特', '正码特_正五特'],
            '正6特': ['正6特', '正六特', '正码特_正六特'],
            '正1': ['正1', '正一', '正码1', '正码_正一'],
            '正2': ['正2', '正二', '正码2', '正码_正二'],
            '正3': ['正3', '正三', '正码3', '正码_正三'],
            '正4': ['正4', '正四', '正码4', '正码_正四'],
            '正5': ['正5', '正五', '正码5', '正码_正五'],
            '正6': ['正6', '正六', '正码6', '正码_正六']
        }
        
        base_directions = {
            '大': ['大', 'big', 'large', 'da'],
            '小': ['小', 'small', 'xiao'],
            '单': ['单', 'odd', 'dan', '奇'],
            '双': ['双', 'even', 'shuang', '偶']
        }
        
        # 检查是否是位置-方向组合
        for position, keywords in lhc_position_map.items():
            for keyword in keywords:
                if keyword in content_lower:
                    for direction, dir_keywords in base_directions.items():
                        for dir_keyword in dir_keywords:
                            if dir_keyword in content_lower:
                                combined_direction = f"{position}-{direction}"
                                directions.add(combined_direction)
                                break
        
        # 如果没有找到组合，查找基础方向
        if not directions:
            for direction, keywords in base_directions.items():
                for keyword in keywords:
                    if (keyword in content_lower and 
                        (len(keyword) > 1 or 
                         (len(keyword) == 1 and 
                          (content_lower == keyword or 
                           f" {keyword} " in f" {content_lower} " or
                           content_lower.startswith(keyword + ' ') or
                           content_lower.endswith(' ' + keyword))))):
                        directions.add(direction)
                        break
        
        return directions

    @staticmethod
    def prioritize_directions(directions, content, play_category):
        """方向优先级排序"""
        if not directions:
            return ""
        
        if len(directions) == 1:
            return directions[0]
        
        content_lower = content.lower()
        play_lower = play_category.lower() if play_category else ""
        
        priority_scores = {}
        
        for direction in directions:
            score = 0
            
            # 精确匹配得分最高
            if direction == content_lower:
                score += 100
            
            # 玩法分类优先级
            if any(word in play_lower for word in ['两面', '和值', '大小单双']):
                score += 50
            
            # 特殊模式优先级
            if '总' in content_lower and '总和' in direction:
                score += 30
            elif '特' in content_lower and '特' in direction:
                score += 30
            
            # 基础方向优先级
            if direction in ['大', '小', '单', '双']:
                score += 20
            
            priority_scores[direction] = score
        
        return max(priority_scores.items(), key=lambda x: x[1])[0]

    @staticmethod
    def extract_position_from_play_category(play_category, lottery_type, config):
        """从玩法分类中提取位置信息"""
        play_str = str(play_category).strip()
        
        if not play_str:
            return '未知位置'
        
        position_keywords = config.position_keywords.get(lottery_type, {})
        
        for position, keywords in position_keywords.items():
            for keyword in keywords:
                if keyword in play_str:
                    return position
        
        # LHC特殊处理
        if lottery_type == 'LHC':
            if '正码特' in play_str or '正特' in play_str:
                return '特码'
            elif '正码' in play_str and '特' not in play_str:
                return '正码'
        
        return '未知位置'

    @staticmethod
    def parse_pk10_vertical_format(content):
        """解析PK10竖线分隔格式"""
        try:
            content_str = str(content).strip()
            bets_by_position = defaultdict(list)
            
            if not content_str:
                return bets_by_position
            
            positions = ['冠军', '亚军', '第三名', '第四名', '第五名', 
                        '第六名', '第七名', '第八名', '第九名', '第十名']
            
            parts = content_str.split('|')
            
            for i, part in enumerate(parts):
                if i < len(positions):
                    position = positions[i]
                    part_clean = part.strip()
                    
                    if not part_clean or part_clean == '_' or part_clean == '':
                        continue
                    
                    numbers = []
                    if ',' in part_clean:
                        number_strs = part_clean.split(',')
                        for num_str in number_strs:
                            num_clean = num_str.strip()
                            if num_clean.isdigit():
                                numbers.append(int(num_clean))
                    else:
                        if part_clean.isdigit():
                            numbers.append(int(part_clean))
                    
                    bets_by_position[position].extend(numbers)
            
            return bets_by_position
        except Exception:
            return defaultdict(list)
    
    @staticmethod
    def parse_3d_vertical_format(content):
        """解析3D竖线分隔格式"""
        try:
            content_str = str(content).strip()
            bets_by_position = defaultdict(list)
            
            if not content_str:
                return bets_by_position
            
            positions = ['百位', '十位', '个位']
            
            parts = content_str.split('|')
            
            for i, part in enumerate(parts):
                if i < len(positions):
                    position = positions[i]
                    part_clean = part.strip()
                    
                    if not part_clean or part_clean == '_' or part_clean == '':
                        continue
                    
                    numbers = []
                    if ',' in part_clean:
                        number_strs = part_clean.split(',')
                        for num_str in number_strs:
                            num_clean = num_str.strip()
                            if num_clean.isdigit():
                                numbers.append(int(num_clean))
                    else:
                        if part_clean.isdigit():
                            numbers.append(int(part_clean))
                    
                    bets_by_position[position].extend(numbers)
            
            return bets_by_position
        except Exception:
            return defaultdict(list)

# ==================== PK拾序列位置检测器 ====================
class PK10SequenceDetector:
    """PK拾序列位置检测器"""
    
    def __init__(self, config=None):
        self.config = config or Config()
        self.content_parser = ContentParser()

        self.play_category_to_positions = {
            '1-5名': ['冠军', '亚军', '第三名', '第四名', '第五名'],
            '6-10名': ['第六名', '第七名', '第八名', '第九名', '第十名'],
            '冠军': ['冠军'],
            '亚军': ['亚军'], 
            '第三名': ['第三名'],
            '第四名': ['第四名'],
            '第五名': ['第五名'],
            '第六名': ['第六名'],
            '第七名': ['第七名'],
            '第八名': ['第八名'],
            '第九名': ['第九名'],
            '第十名': ['第十名'],
            '定位胆': ['冠军', '亚军', '第三名', '第四名', '第五名', 
                     '第六名', '第七名', '第八名', '第九名', '第十名']
        }
        
        self.direction_mapping = {
            '大': ['大', 'big', 'large', 'da'],
            '小': ['小', 'small', 'xiao'], 
            '单': ['单', 'odd', 'dan', '奇'],
            '双': ['双', 'even', 'shuang', '偶'],
            '龙': ['龙', 'long', 'dragon'],
            '虎': ['虎', 'hu', 'tiger']
        }
        
        self.pk10_positions = [
            '冠军', '亚军', '第三名', '第四名', '第五名',
            '第六名', '第七名', '第八名', '第九名', '第十名'
        ]
        
    def extract_pk10_bet_content(self, content, play_category):
        """提取PK10投注内容"""
        try:
            if pd.isna(content):
                return None
            
            content_str = str(content).strip()
            
            if ',' in content_str and any(pos in content_str for pos in self.pk10_positions):
                return self._parse_comma_separated_format(content_str)
            
            directions = self.content_parser.enhanced_extract_directions(content_str, self.config)
            if directions:
                return directions[0]
            
            return None
            
        except Exception as e:
            logger.warning(f"PK10内容提取失败: {content}, 错误: {e}")
            return None
    
    def _parse_comma_separated_format(self, content):
        """解析逗号分隔的位置-方向格式"""
        try:
            items = content.split(',')
            directions_found = set()
            
            for item in items:
                item_clean = item.strip()
                if '-' in item_clean:
                    direction_part = item_clean.split('-')[-1].strip()
                    
                    for direction, keywords in self.direction_mapping.items():
                        for keyword in keywords:
                            if direction_part == keyword or direction_part in keyword:
                                directions_found.add(direction)
                                break
            
            if len(directions_found) == 1:
                return list(directions_found)[0]
            
            return None
            
        except Exception as e:
            logger.debug(f"逗号分隔格式解析失败: {content}, 错误: {e}")
            return None
    
    def get_positions_from_play_category(self, play_category):
        """从玩法分类获取对应的位置列表"""
        play_str = str(play_category).strip()
        return self.play_category_to_positions.get(play_str, [])
    
    def _detect_incomplete_position_collaboration(self, period_data, period):
        """检测不完整位置的协作模式"""
        patterns = []
        
        # 检查1-5名投注
        play_1_5 = period_data[period_data['玩法分类'] == '1-5名']
        play_6_10 = period_data[period_data['玩法分类'] == '6-10名']
        
        if len(play_1_5) == 0 or len(play_6_10) == 0:
            return patterns
        
        # 分析每个账户的投注内容
        account_bets = defaultdict(lambda: {'1_5_bets': [], '6_10_bets': []})
        
        for _, row in period_data.iterrows():
            account = row['会员账号']
            play_category = row['玩法分类']
            content = row['内容']
            direction = row.get('投注方向', '')
            amount = row.get('投注金额', 0)
            
            if play_category == '1-5名':
                account_bets[account]['1_5_bets'].append({
                    'content': content,
                    'direction': direction,
                    'amount': amount
                })
            elif play_category == '6-10名':
                account_bets[account]['6_10_bets'].append({
                    'content': content,
                    'direction': direction,
                    'amount': amount
                })
        
        # 查找可能的不完整协作
        accounts = list(account_bets.keys())
        for i in range(len(accounts)):
            for j in range(i+1, len(accounts)):
                acc1 = accounts[i]
                acc2 = accounts[j]
                
                # 检查是否是典型的"你投1-5名，我投6-10名"模式
                bets1 = account_bets[acc1]
                bets2 = account_bets[acc2]
                
                # 情况1：acc1投1-5名，acc2投6-10名
                if len(bets1['1_5_bets']) > 0 and len(bets2['6_10_bets']) > 0:
                    # 检查投注方向是否相同
                    direction1 = bets1['1_5_bets'][0]['direction']
                    direction2 = bets2['6_10_bets'][0]['direction']
                    
                    if direction1 and direction2 and direction1 == direction2:
                        # 检查投注内容是否匹配
                        content1 = bets1['1_5_bets'][0]['content']
                        content2 = bets2['6_10_bets'][0]['content']
                        
                        # 检查是否为简单投注（如只投冠军）
                        if '冠军-' in content1 and len(content1.split(',')) == 1:
                            pattern_type = '冠军单点协作'
                        else:
                            # 分析投注的详细位置
                            positions_1_5 = self._extract_positions_from_content(content1)
                            positions_6_10 = self._extract_positions_from_content(content2)
                            
                            pattern_type = f'部分位置协作({len(positions_1_5)}个1-5名, {len(positions_6_10)}个6-10名)'
                        
                        amounts = [
                            sum(b['amount'] for b in bets1['1_5_bets']),
                            sum(b['amount'] for b in bets2['6_10_bets'])
                        ]
                        
                        patterns.append({
                            '期号': period,
                            '彩种': 'PK10',
                            '彩种类型': 'PK10',
                            '账户组': [acc1, acc2],
                            '方向组': [direction1, direction2],
                            '金额组': amounts,
                            '总金额': sum(amounts),
                            '相似度': 1.0,
                            '账户数量': 2,
                            '模式': f'PK10-{pattern_type}-{direction1.replace("数字-", "")}',
                            '对立类型': f'{pattern_type}-{direction1}',
                            '检测类型': 'PK10序列位置'
                        })
                
                # 情况2：acc2投1-5名，acc1投6-10名
                if len(bets2['1_5_bets']) > 0 and len(bets1['6_10_bets']) > 0:
                    # 类似上面的逻辑，只是账户顺序相反
                    direction1 = bets2['1_5_bets'][0]['direction']
                    direction2 = bets1['6_10_bets'][0]['direction']
                    
                    if direction1 and direction2 and direction1 == direction2:
                        amounts = [
                            sum(b['amount'] for b in bets2['1_5_bets']),
                            sum(b['amount'] for b in bets1['6_10_bets'])
                        ]
                        
                        patterns.append({
                            '期号': period,
                            '彩种': 'PK10',
                            '彩种类型': 'PK10',
                            '账户组': [acc2, acc1],
                            '方向组': [direction1, direction2],
                            '金额组': amounts,
                            '总金额': sum(amounts),
                            '相似度': 1.0,
                            '账户数量': 2,
                            '模式': f'PK10-部分位置协作-{direction1.replace("数字-", "")}',
                            '对立类型': f'部分位置协作-{direction1}',
                            '检测类型': 'PK10序列位置'
                        })
        
        return patterns
    
    def _extract_positions_from_content(self, content):
        """从内容中提取位置信息"""
        positions = []
        content_str = str(content)
        
        position_keywords = ['冠军', '亚军', '第三名', '第四名', '第五名',
                            '第六名', '第七名', '第八名', '第九名', '第十名']
        
        for position in position_keywords:
            if position in content_str:
                positions.append(position)
        
        return positions
    
    def detect_sequence_coverage(self, df_pk10):
        """检测序列覆盖模式"""
        sequence_patterns = []
        
        period_groups = df_pk10.groupby('期号')
        
        for period, period_data in period_groups:
            position_account_content = defaultdict(lambda: defaultdict(list))
            
            for _, row in period_data.iterrows():
                account = row['会员账号']
                play_category = row.get('玩法分类', '')
                content = row['内容']
                amount = row.get('投注金额', 0)
                
                positions_from_play = self.get_positions_from_play_category(play_category)
                positions_from_content = self._extract_positions_from_content(content)
                
                all_positions = list(set(positions_from_play + positions_from_content))
                if not all_positions:
                    position = self.content_parser.extract_position_from_play_category(
                        play_category, 'PK10', self.config
                    )
                    if position in self.pk10_positions:
                        all_positions = [position]
                
                bet_content = self.extract_pk10_bet_content(content, play_category)
                if bet_content is None:
                    continue
                
                for position in all_positions:
                    if position in self.pk10_positions:
                        position_account_content[position][account].append({
                            'content': bet_content,
                            'amount': amount,
                            'original_content': content,
                            'play_category': play_category,
                            'positions_covered': all_positions
                        })
            
            patterns = self._find_sequence_coverage_patterns(
                position_account_content, period
            )
            sequence_patterns.extend(patterns)
        
        return sequence_patterns
    
    def _find_sequence_coverage_patterns(self, position_account_content, period):
        """查找序列覆盖模式"""
        patterns = []
        
        all_accounts = set()
        account_bet_contents = defaultdict(set)
        
        for position, account_data in position_account_content.items():
            for account, bets in account_data.items():
                all_accounts.add(account)
                for bet in bets:
                    account_bet_contents[account].add(str(bet['content']))
        
        common_content_groups = defaultdict(list)
        
        for account, contents in account_bet_contents.items():
            for content in contents:
                common_content_groups[content].append(account)
        
        for bet_content, accounts in common_content_groups.items():
            if len(accounts) < 2:
                continue
            
            if len(accounts) >= 2:
                for account_group in combinations(accounts, 2):
                    coverage_result = self._check_position_coverage(
                        position_account_content, list(account_group), bet_content
                    )
                    if coverage_result['covered']:
                        pattern = self._create_sequence_pattern(
                            period, list(account_group), bet_content, coverage_result
                        )
                        patterns.append(pattern)
            
            if len(accounts) >= 3:
                for account_group in combinations(accounts, 3):
                    coverage_result = self._check_position_coverage(
                        position_account_content, list(account_group), bet_content
                    )
                    if coverage_result['covered']:
                        pattern = self._create_sequence_pattern(
                            period, list(account_group), bet_content, coverage_result
                        )
                        patterns.append(pattern)
        
        return patterns
    
    def _check_position_coverage(self, position_account_content, accounts, target_content):
        """检查账户组是否覆盖了十个位置且投注内容相同"""
        covered_positions = set()
        position_details = {}
        total_amount = 0
        
        for position in self.pk10_positions:
            if position not in position_account_content:
                continue
            
            position_covered = False
            position_accounts = []
            position_amounts = []
            
            for account in accounts:
                if account in position_account_content[position]:
                    account_bets = position_account_content[position][account]
                    for bet in account_bets:
                        bet_content_str = str(bet['content'])
                        if bet_content_str == target_content:
                            position_covered = True
                            position_accounts.append(account)
                            position_amounts.append(bet['amount'])
                            total_amount += bet['amount']
                            break
            
            if position_covered:
                covered_positions.add(position)
                position_details[position] = {
                    'accounts': position_accounts,
                    'amounts': position_amounts
                }
        
        return {
            'covered': len(covered_positions) == len(self.pk10_positions),
            'covered_positions': covered_positions,
            'position_details': position_details,
            'total_amount': total_amount
        }
    
    def _create_sequence_pattern(self, period, accounts, bet_content, coverage_result):
        """创建序列覆盖模式记录"""
        coverage_ratio = len(coverage_result['covered_positions']) / len(self.pk10_positions)
        
        detailed_records = []
        for position in self.pk10_positions:
            if position in coverage_result['position_details']:
                details = coverage_result['position_details'][position]
                record = {
                    'position': position,
                    'accounts': details['accounts'],
                    'amounts': details['amounts'],
                    'bet_content': bet_content
                }
                detailed_records.append(record)
        
        account_count = len(accounts)
        if account_count == 2:
            pattern_desc = f'PK10十位置全覆盖-{bet_content}(2账户协作)'
        elif account_count == 3:
            pattern_desc = f'PK10十位置全覆盖-{bet_content}(3账户协作)'
        else:
            pattern_desc = f'PK10十位置全覆盖-{bet_content}({account_count}账户协作)'
        
        return {
            '期号': period,
            '彩种': 'PK10',
            '彩种类型': 'PK10',
            '账户组': accounts,
            '账户数量': account_count,
            '投注内容': bet_content,
            '覆盖位置数': len(coverage_result['covered_positions']),
            '总位置数': len(self.pk10_positions),
            '覆盖度': coverage_ratio,
            '总投注金额': coverage_result['total_amount'],
            '位置详情': detailed_records,
            '模式类型': '序列覆盖',
            '模式描述': pattern_desc
        }

    def _extract_direction_from_data(self, data):
        """从数据中提取主要投注方向"""
        try:
            if len(data) == 0:
                return None
            
            if '投注方向' not in data.columns:
                return None
            
            direction = data.iloc[0]['投注方向']
            return direction
            
        except Exception as e:
            return None
    
    def _extract_number_from_content(self, content):
        """从内容中提取数字"""
        try:
            if pd.isna(content):
                return None
            
            content_str = str(content).strip()
            
            if '-' in content_str:
                parts = content_str.split('-')
                if len(parts) >= 2:
                    number_part = parts[-1].strip()
                    if number_part.isdigit():
                        return number_part
            
            numbers = re.findall(r'\d+', content_str)
            if numbers:
                return numbers[0]
            
            return None
        except:
            return None
    
    def _parse_pk10_content_enhanced(self, data):
        """增强版PK10内容解析"""
        if len(data) == 0:
            return None
        
        sample_row = data.iloc[0]
        content = sample_row['内容']
        
        parsed_content = self.content_parser.parse_complex_content(content, '')
        
        content_type = parsed_content.get('type', 'unknown')
        
        if content_type == 'number':
            return f"数字-{parsed_content['value']}"
        elif content_type == 'single_position':
            value = parsed_content['value']
            if value.isdigit():
                return f"数字-{value}"
            return value
        elif content_type == 'multiple_positions':
            value = parsed_content['value']
            if value.isdigit():
                return f"数字-{value}"
            return value
        elif content_type == 'raw':
            directions = self.content_parser.enhanced_extract_directions(content, self.config)
            if directions:
                return directions[0]
            
            numbers = self.content_parser.extract_all_numbers(content)
            if numbers:
                return f"数字-{numbers[0]}"
            
            return content
        else:
            return str(content)

# ==================== 对刷检测器类 ====================
class WashTradeDetector:
    def __init__(self, config=None):
        self.config = config or Config()
        self.data_processor = DataProcessor()
        self.lottery_identifier = LotteryIdentifier()
        self.play_normalizer = PlayCategoryNormalizer()
        self.content_parser = ContentParser()
        
        self.data_processed = False
        self.df_valid = None
        self.export_data = []
        self.pk10_sequence_detector = PK10SequenceDetector(config)
        
        self.account_total_periods_by_lottery = defaultdict(dict)
        self.account_record_stats_by_lottery = defaultdict(dict)
        self.performance_stats = {}

    def filter_accounts_by_amount_balance(self, account_group, directions, amounts):
        """根据组内金额平衡性过滤账户 - 确保正确过滤"""
        if not self.config.amount_threshold['enable_threshold_filter']:
            return account_group, directions, amounts
        
        if not amounts or len(amounts) < 2:
            return account_group, directions, amounts
        
        # 计算最大最小金额比例
        max_amount = max(amounts)
        min_amount = min(amounts)
        
        # 防止除零
        if min_amount == 0:
            return account_group, directions, amounts
        
        amount_ratio = max_amount / min_amount
        
        max_allowed_ratio = self.config.amount_threshold['max_amount_ratio']
        
        # 如果金额比例超过阈值，直接过滤掉这个组合
        if amount_ratio > max_allowed_ratio:
            logger.info(f"金额平衡过滤: 账户组 {account_group} 金额比例 {amount_ratio:.1f}倍 > 阈值 {max_allowed_ratio}倍，过滤")
            logger.info(f"原始金额: {amounts}")
            return [], [], []
        
        logger.info(f"金额平衡检查通过: 账户组 {account_group} 金额比例 {amount_ratio:.1f}倍 <= 阈值 {max_allowed_ratio}倍")
        return account_group, directions, amounts

    def upload_and_process(self, uploaded_file):
        """上传并处理文件"""
        try:
            if uploaded_file is None:
                st.error("❌ 没有上传文件")
                return None, None
            
            filename = uploaded_file.name
            logger.info(f"✅ 已上传文件: {filename}")
            
            if not any(filename.endswith(ext) for ext in self.config.supported_file_types):
                st.error(f"❌ 不支持的文件类型: {filename}")
                return None, None
            
            with st.spinner("🔄 正在清洗数据..."):
                df_clean = self.data_processor.clean_data(uploaded_file)
            
            if df_clean is not None and len(df_clean) > 0:
                df_enhanced = self.enhance_data_processing(df_clean)
                return df_enhanced, filename
            else:
                return None, None
            
        except Exception as e:
            logger.error(f"文件处理失败: {str(e)}")
            st.error(f"文件处理失败: {str(e)}")
            return None, None
    
    def enhance_data_processing(self, df_clean):
        """数据处理流程"""
        try:
            if '彩种' in df_clean.columns:
                df_clean['原始彩种'] = df_clean['彩种']
                df_clean['彩种类型'] = df_clean['彩种'].apply(self.lottery_identifier.identify_lottery_type)
            
            if '玩法' in df_clean.columns:
                df_clean['玩法分类'] = df_clean['玩法'].apply(self.play_normalizer.normalize_category)
            else:
                df_clean['玩法分类'] = ''
            
            df_clean['投注金额'] = df_clean['金额'].apply(
                lambda x: self.extract_bet_amount_safe(str(x))
            )
            
            df_clean['投注方向'] = df_clean.apply(
                lambda row: self.enhanced_extract_direction_with_position(
                    row['内容'], 
                    row.get('玩法分类', ''), 
                    row.get('彩种类型', '未知')
                ), 
                axis=1
            )
            
            df_valid = df_clean[
                (df_clean['投注方向'] != '') & 
                (df_clean['投注金额'] >= self.config.min_amount)
            ].copy()
            
            self.data_processed = True
            self.df_valid = df_valid
            
            self.calculate_account_total_periods_by_lottery(df_valid)
            
            return df_valid
                
        except Exception as e:
            logger.error(f"数据处理增强失败: {str(e)}")
            st.error(f"数据处理增强失败: {str(e)}")
            return pd.DataFrame()

    def extract_bet_amount_safe(self, amount_text):
        """安全提取投注金额"""
        try:
            if pd.isna(amount_text):
                return 0
            
            text = str(amount_text).strip()
            
            # 处理特殊格式：投注：xx抵用：xx
            if '投注：' in text and '抵用：' in text:
                try:
                    bet_part = text.split('投注：')[1].split('抵用：')[0].strip()
                    amount = float(bet_part.replace(',', ''))
                    if amount >= self.config.min_amount:
                        return amount
                except (ValueError, IndexError):
                    pass
            
            # 处理简化格式：投注：xx
            if text.startswith('投注：'):
                try:
                    bet_part = text.replace('投注：', '').strip()
                    bet_part_clean = re.split(r'[^\d.]', bet_part)[0]
                    amount = float(bet_part_clean)
                    if amount >= self.config.min_amount:
                        return amount
                except (ValueError, IndexError):
                    pass
            
            # 处理英文冒号格式
            if '投注:' in text:
                try:
                    bet_part = text.split('投注:')[1].split()[0].strip()
                    amount = float(bet_part.replace(',', ''))
                    if amount >= self.config.min_amount:
                        return amount
                except (ValueError, IndexError):
                    pass
            
            # 处理科学计数法
            if 'E' in text or 'e' in text:
                try:
                    amount = float(text)
                    if amount >= self.config.min_amount:
                        return amount
                except:
                    pass
            
            # 尝试提取纯数字
            try:
                cleaned_text = re.sub(r'[^\d.-]', '', text)
                if cleaned_text and cleaned_text != '-':
                    amount = float(cleaned_text)
                    if amount >= self.config.min_amount:
                        return amount
            except:
                pass
            
            # 使用正则表达式模式匹配
            patterns = [
                r'投注[:：]?\s*([-]?\d+[,，]?\d*\.?\d*)',
                r'下注[:：]?\s*([-]?\d+[,，]?\d*\.?\d*)',
                r'金额[:：]?\s*([-]?\d+[,，]?\d*\.?\d*)',
                r'总额[:：]?\s*([-]?\d+[,，]?\d*\.?\d*)',
                r'([-]?\d+[,，]?\d*\.?\d*)\s*元',
                r'￥\s*([-]?\d+[,，]?\d*\.?\d*)',
                r'¥\s*([-]?\d+[,，]?\d*\.?\d*)',
                r'[\$￥¥]?\s*([-]?\d+[,，]?\d*\.?\d+)',
                r'([-]?\d+[,，]?\d*\.?\d+)',
            ]
            
            for pattern in patterns:
                match = re.search(pattern, text)
                if match:
                    amount_str = match.group(1).replace(',', '').replace('，', '').replace(' ', '')
                    try:
                        amount = float(amount_str)
                        if amount >= self.config.min_amount:
                            return amount
                    except:
                        continue
            
            return 0
                
        except Exception:
            return 0
    
    def enhanced_extract_direction_with_position(self, content, play_category, lottery_type):
        """全面修复方向提取 - 支持所有格式"""
        try:
            if pd.isna(content):
                return ""
            
            content_str = str(content).strip()
            
            # 如果是PK10类型
            if lottery_type == 'PK10':
                # 格式1: 冠军-01,04,05（多个数字）
                if '-' in content_str and ',' in content_str:
                    # 检查是否是位置-数字格式
                    parts = content_str.split('-', 1)
                    if len(parts) >= 2:
                        number_part = parts[1].strip()
                        # 提取所有数字
                        numbers = re.findall(r'\b\d{1,2}\b', number_part)
                        if numbers:
                            unique_numbers = sorted(set(numbers))
                            if len(unique_numbers) > 1:
                                return f"多数字-{','.join(unique_numbers)}"
                            elif len(unique_numbers) == 1:
                                return f"数字-{unique_numbers[0]}"
                
                # 格式2: 冠军-双（方向）
                if '-' in content_str:
                    parts = content_str.split('-', 1)
                    if len(parts) >= 2:
                        value_part = parts[1].strip()
                        # 检查是否是方向（大小单双等）
                        directions = self.content_parser.enhanced_extract_directions(value_part, self.config)
                        if directions:
                            return directions[0]  # 取第一个方向
            
            # 通用数字提取
            numbers = re.findall(r'\b\d{1,2}\b', content_str)
            if numbers:
                unique_numbers = sorted(set(numbers))
                if len(unique_numbers) > 1:
                    return f"多数字-{','.join(unique_numbers)}"
                elif len(unique_numbers) == 1:
                    return f"数字-{unique_numbers[0]}"
            
            # 通用方向提取
            directions = self.content_parser.enhanced_extract_directions(content_str, self.config)
            if directions:
                return directions[0]  # 取第一个方向
            
            return ""
                
        except Exception as e:
            logger.warning(f"方向提取失败: {content}, 错误: {e}")
            return ""
    
    def _select_primary_direction(self, directions, content):
        """选择主要方向"""
        if not directions:
            return ""
        
        if len(directions) == 1:
            return directions[0]
        
        content_str = str(content)
        
        priority_rules = [
            lambda d: any(keyword in content_str for keyword in ['总和', '总']) and d in directions,
            lambda d: '特' in content_str and d in directions,
            lambda d: any(keyword in content_str for keyword in ['和值', '和']) and d in directions,
            lambda d: '两面' in content_str and d in directions,
            lambda d: d in directions
        ]
        
        for rule in priority_rules:
            matching_directions = [d for d in directions if rule(d)]
            if matching_directions:
                return matching_directions[0]
        
        return directions[0]
    
    def _extract_position_from_content(self, content, lottery_type):
        """从内容中提取位置信息"""
        content_str = str(content).strip()
        
        position_keywords = self.config.position_keywords.get(lottery_type, {})
        
        for position, keywords in position_keywords.items():
            for keyword in keywords:
                if keyword in content_str:
                    return position
        
        if '|' in content_str:
            if lottery_type == 'PK10':
                bets_by_position = self.content_parser.parse_pk10_vertical_format(content_str)
                for position in bets_by_position:
                    if bets_by_position[position]:
                        return position
            elif lottery_type == '3D':
                bets_by_position = self.content_parser.parse_3d_vertical_format(content_str)
                for position in bets_by_position:
                    if bets_by_position[position]:
                        return position
        
        return '未知位置'
    
    def calculate_account_total_periods_by_lottery(self, df):
        """修复账户期数统计方法"""
        self.account_total_periods_by_lottery = defaultdict(dict)
        self.account_record_stats_by_lottery = defaultdict(dict)
        
        data_source = self.df_valid if hasattr(self, 'df_valid') and self.df_valid is not None else df
        
        lottery_col = '彩种'
        
        for lottery in data_source[lottery_col].unique():
            df_lottery = data_source[data_source[lottery_col] == lottery]
            
            period_counts = df_lottery.groupby('会员账号')['期号'].nunique().to_dict()
            self.account_total_periods_by_lottery[lottery] = period_counts
            
            record_counts = df_lottery.groupby('会员账号').size().to_dict()
            self.account_record_stats_by_lottery[lottery] = record_counts
    
    def detect_all_wash_trades(self):
        """修复的主检测方法"""
        if not self.data_processed or self.df_valid is None or len(self.df_valid) == 0:
            st.error("❌ 没有有效数据可用于检测")
            return []
        
        df_filtered = self.exclude_multi_direction_accounts(self.df_valid)
        
        if len(df_filtered) == 0:
            st.error("❌ 过滤后无有效数据")
            return []
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        all_patterns = []
        total_steps = self.config.max_accounts_in_group + 1
        
        for account_count in range(2, self.config.max_accounts_in_group + 1):
            status_text.text(f"🔍 检测{account_count}个账户对刷模式...")
            patterns = self.detect_n_account_patterns_optimized(df_filtered, account_count)
            all_patterns.extend(patterns)
            
            progress = (account_count - 1) / total_steps
            progress_bar.progress(progress)
        
        status_text.text(f"🔍 检测PK10序列位置模式...")
        pk10_patterns = self.detect_pk10_sequence_patterns(df_filtered)
        all_patterns.extend(pk10_patterns)
        
        progress_bar.progress(1.0)
        status_text.empty()
        
        return all_patterns
    
    def detect_n_account_patterns_optimized(self, df_filtered, n_accounts):
        """N个账户对刷模式检测"""
        wash_records = []
        
        period_groups = df_filtered.groupby(['期号', '原始彩种'])
        
        valid_direction_combinations = self._get_valid_direction_combinations(n_accounts)
        
        batch_size = 100
        period_keys = list(period_groups.groups.keys())
        
        for i in range(0, len(period_keys), batch_size):
            batch_keys = period_keys[i:i+batch_size]
            
            for period_key in batch_keys:
                period_data = period_groups.get_group(period_key)
                period_accounts = period_data['会员账号'].unique()
                
                if len(period_accounts) < n_accounts:
                    continue
                
                batch_patterns = self._detect_combinations_for_period(
                    period_data, period_accounts, n_accounts, valid_direction_combinations
                )
                wash_records.extend(batch_patterns)
        
        return self.find_continuous_patterns_optimized(wash_records)

    def detect_pk10_sequence_patterns(self, df_filtered):
        """PK10序列位置模式检测 - 保持所有检测逻辑"""
        try:
            if hasattr(self, 'df_valid') and self.df_valid is not None:
                df_pk10 = self.df_valid[
                    (self.df_valid['彩种类型'] == 'PK10') & 
                    (self.df_valid['投注金额'] >= self.config.min_amount)
                ].copy()
            else:
                df_pk10 = df_filtered[
                    (df_filtered['彩种类型'] == 'PK10') & 
                    (df_filtered['投注金额'] >= self.config.min_amount)
                ].copy()
            
            if len(df_pk10) == 0:
                return []
            
            sequence_patterns = []
            period_groups = df_pk10.groupby('期号')
            
            for period, period_data in period_groups:
                if len(period_data) > 0:
                    if '原始彩种' in period_data.columns:
                        specific_lottery = period_data['原始彩种'].iloc[0]
                    else:
                        specific_lottery = period_data['彩种'].iloc[0]
                else:
                    specific_lottery = 'PK10'
                
                # 调用所有检测方法
                patterns_1 = self._detect_1_5_6_10_collaboration(period_data, period, specific_lottery)
                sequence_patterns.extend(patterns_1)
                
                patterns_2 = self._detect_single_position_full_coverage(period_data, period, specific_lottery)
                sequence_patterns.extend(patterns_2)
                
                patterns_3 = self._detect_vertical_format_collaboration(period_data, period, specific_lottery)
                sequence_patterns.extend(patterns_3)
                
                # 新增：检测任意位置分配组合
                patterns_4 = self._detect_arbitrary_position_coverage(period_data, period, specific_lottery)
                sequence_patterns.extend(patterns_4)
            
            # 使用修复后的连续模式检测方法，它会进行适当的去重
            continuous_patterns = self.find_continuous_patterns_optimized(sequence_patterns)
            
            return continuous_patterns
                    
        except Exception as e:
            logger.error(f"PK10序列检测失败: {str(e)}")
            return []
    
    def _get_valid_direction_combinations(self, n_accounts):
        """有效方向组合生成"""
        valid_combinations = []
        
        for opposites in self.config.opposite_groups:
            opposite_list = list(opposites)
            
            if n_accounts == 2:
                if len(opposite_list) == 2:
                    dir1, dir2 = opposite_list
                    valid_combinations.append({
                        'directions': [dir1, dir2],
                        'dir1_count': 1,
                        'dir2_count': 1,
                        'opposite_type': f"{dir1}-{dir2}",
                        'combination_type': 'basic'
                    })
            else:
                for i in range(1, n_accounts):
                    j = n_accounts - i
                    if len(opposite_list) == 2:
                        dir1, dir2 = opposite_list
                        valid_combinations.append({
                            'directions': [dir1] * i + [dir2] * j,
                            'dir1_count': i,
                            'dir2_count': j,
                            'opposite_type': f"{dir1}-{dir2}",
                            'combination_type': 'basic'
                        })
        
        multi_number_combinations = [
            ['多数字-01,04,05', '多数字-01,04,05'],
            ['多数字-01,04,05', '多数字-01,04,05'],
        ]
        
        for combo in multi_number_combinations:
            if n_accounts == len(combo):
                valid_combinations.append({
                    'directions': combo,
                    'dir1_count': n_accounts,
                    'dir2_count': 0,
                    'opposite_type': f"协作覆盖-多数字",
                    'combination_type': 'multi_number'
                })
        
        return valid_combinations
    
    def _detect_combinations_for_period(self, period_data, period_accounts, n_accounts, valid_combinations):
        """为单个期号检测组合"""
        patterns = []
        detected_combinations = set()
        
        # 确保lottery_type有默认值
        lottery_type = '未知'
        
        # 尝试从不同列获取彩种类型
        if len(period_data) > 0:
            if '彩种类型' in period_data.columns:
                lottery_type = period_data['彩种类型'].iloc[0]
            elif '原始彩种' in period_data.columns:
                # 从原始彩种推断类型
                lottery_name = period_data['原始彩种'].iloc[0]
                lottery_type = self.lottery_identifier.identify_lottery_type(lottery_name)
            elif '彩种' in period_data.columns:
                lottery_name = period_data['彩种'].iloc[0]
                lottery_type = self.lottery_identifier.identify_lottery_type(lottery_name)
        
        lottery = period_data['原始彩种'].iloc[0] if '原始彩种' in period_data.columns else period_data['彩种'].iloc[0]
        
        current_period = period_data['期号'].iloc[0]
        
        # 修复点：同一账户同一方向的多笔投注金额合并
        # 使用嵌套的defaultdict来合并同一账户同一方向的金额
        account_direction_amounts = defaultdict(lambda: defaultdict(float))
        
        for _, row in period_data.iterrows():
            account = row['会员账号']
            direction = row['投注方向']
            amount = row['投注金额']
            
            if direction:  # 只处理有方向的记录
                # 累加同一账户同一方向的金额
                account_direction_amounts[account][direction] += amount
        
        # 将合并后的数据转换回原来的数据结构格式
        account_info = {}
        for account, direction_amounts in account_direction_amounts.items():
            # 每个账户可能有多个方向，但我们只取一个（因为已过滤多方向账户）
            if direction_amounts:
                # 取第一个方向（因为我们过滤了多方向账户）
                direction = list(direction_amounts.keys())[0]
                total_amount = direction_amounts[direction]
                account_info[account] = [{
                    'direction': direction,
                    'amount': total_amount
                }]
        
        for account_group in combinations(period_accounts, n_accounts):
            if not self._check_account_period_difference(account_group, lottery):
                continue
            
            group_directions = []
            group_amounts = []
            
            for account in account_group:
                if account in account_info and account_info[account]:
                    first_bet = account_info[account][0]
                    group_directions.append(first_bet['direction'])
                    group_amounts.append(first_bet['amount'])
            
            if len(group_directions) != n_accounts:
                continue
            
            filtered_account_group, filtered_directions, filtered_amounts = self.filter_accounts_by_amount_balance(
                account_group, group_directions, group_amounts
            )
            
            if len(filtered_account_group) < 2:
                continue
            
            account_group = filtered_account_group
            group_directions = filtered_directions
            group_amounts = filtered_amounts
            n_accounts = len(account_group)

            combination_key = (
                tuple(sorted(account_group)), 
                tuple(sorted(group_directions)),
                tuple(sorted(group_amounts))
            )
            
            if combination_key in detected_combinations:
                continue
            
            for combo in valid_combinations:
                target_directions = combo['directions']
                
                actual_directions_sorted = sorted(group_directions)
                target_directions_sorted = sorted(target_directions)
                
                if actual_directions_sorted == target_directions_sorted:
                    detected_combinations.add(combination_key)
                    
                    dir1_total = 0
                    dir2_total = 0
                    dir1 = combo['directions'][0]
                    
                    for direction, amount in zip(group_directions, group_amounts):
                        if direction == dir1:
                            dir1_total += amount
                        else:
                            dir2_total += amount
                    
                    similarity_threshold = self.config.account_count_similarity_thresholds.get(
                        n_accounts, self.config.amount_similarity_threshold
                    )
                    
                    if dir1_total > 0 and dir2_total > 0:
                        similarity = min(dir1_total, dir2_total) / max(dir1_total, dir2_total)
                        
                        if similarity >= similarity_threshold:
                            # 使用已经定义好的lottery_type
                            if ' vs ' in combo['opposite_type']:
                                pattern_parts = combo['opposite_type'].split(' vs ')
                                if len(pattern_parts) == 2:
                                    dir1_part = pattern_parts[0].split('-')
                                    dir2_part = pattern_parts[1].split('-')
                                    if len(dir1_part) == 2 and len(dir2_part) == 2:
                                        pattern_str = f"{dir1_part[0]}-{dir1_part[1]}({combo['dir1_count']}个) vs {dir2_part[0]}-{dir2_part[1]}({combo['dir2_count']}个)"
                                    else:
                                        pattern_str = f"{pattern_parts[0]}({combo['dir1_count']}个) vs {pattern_parts[1]}({combo['dir2_count']}个)"
                                else:
                                    pattern_str = combo['opposite_type']
                            else:
                                opposite_parts = combo['opposite_type'].split('-')
                                if len(opposite_parts) == 2:
                                    pattern_str = f"{opposite_parts[0]}({combo['dir1_count']}个) vs {opposite_parts[1]}({combo['dir2_count']}个)"
                                else:
                                    pattern_str = combo['opposite_type']
                            
                            record = {
                                '期号': period_data['期号'].iloc[0],
                                '彩种': lottery,
                                '彩种类型': lottery_type,
                                '账户组': list(account_group),
                                '方向组': group_directions,
                                '金额组': group_amounts,
                                '总金额': dir1_total + dir2_total,
                                '相似度': similarity,
                                '账户数量': n_accounts,
                                '模式': pattern_str,
                                '对立类型': combo['opposite_type']
                            }
                            
                            patterns.append(record)
                            break
        
        return patterns
    
    def _check_account_period_difference(self, account_group, lottery):
        """检查账户组内账户的总投注期数差异是否在阈值内"""
        if lottery not in self.account_total_periods_by_lottery:
            return True
        
        total_periods_stats = self.account_total_periods_by_lottery[lottery]
        
        account_periods = []
        for account in account_group:
            if account in total_periods_stats:
                account_periods.append(total_periods_stats[account])
            else:
                return True
        
        if len(account_periods) < 2:
            return True
        
        max_period = max(account_periods)
        min_period = min(account_periods)
        period_diff = max_period - min_period
        
        if period_diff > self.config.account_period_diff_threshold:
            return False
        
        return True
    
    def find_continuous_patterns_optimized(self, wash_records):
        """连续对刷模式检测 - 修复过度过滤问题"""
        if not wash_records:
            return []
        
        account_group_patterns = defaultdict(list)
        
        # 简单的去重：确保同一期号、同一账户组、同一方向不会重复
        seen_keys = set()
        unique_records = []
        
        for record in wash_records:
            # 创建唯一标识键
            key = (
                record['期号'],
                tuple(sorted(record['账户组'])),
                tuple(sorted(record['方向组']))
            )
            
            if key not in seen_keys:
                seen_keys.add(key)
                unique_records.append(record)
        
        # 使用去重后的记录进行分组
        for record in unique_records:
            # 不再过度过滤PK10序列位置检测
            # 即使是普通协作，也允许显示
            account_group_key = (tuple(sorted(record['账户组'])), record['彩种'])
            account_group_patterns[account_group_key].append(record)
        
        continuous_patterns = []
        
        for account_group_key, records in account_group_patterns.items():
            # 对每个组的记录按期号排序
            sorted_records = sorted(records, key=lambda x: x['期号'])
            
            if isinstance(account_group_key, tuple) and len(account_group_key) > 0:
                if isinstance(account_group_key[0], tuple):
                    account_group = list(account_group_key[0])
                    lottery = account_group_key[1]
                else:
                    account_group = list(account_group_key)
                    lottery = records[0]['彩种'] if records else '未知'
            else:
                continue
            
            # 根据检测类型设置不同的最小期数要求
            if records and '检测类型' in records[0]:
                if records[0]['检测类型'] == 'PK10序列位置':
                    required_min_periods = 3  # PK10完整协作要求至少3期
                else:
                    required_min_periods = self.get_required_min_periods(account_group, lottery)
            else:
                required_min_periods = self.get_required_min_periods(account_group, lottery)
            
            if len(sorted_records) >= required_min_periods:
                # 确保详细记录也是唯一的（按期号去重）
                seen_periods = set()
                unique_detailed_records = []
                
                for record in sorted_records:
                    period = record['期号']
                    if period not in seen_periods:
                        seen_periods.add(period)
                        unique_detailed_records.append(record)
                
                total_investment = sum(r['总金额'] for r in unique_detailed_records)
                similarities = [r['相似度'] for r in unique_detailed_records if '相似度' in r]
                avg_similarity = np.mean(similarities) if similarities else 1.0
                
                opposite_type_counts = defaultdict(int)
                for record in unique_detailed_records:
                    opposite_type = record.get('对立类型', '协作模式')
                    opposite_type_counts[opposite_type] += 1
                
                pattern_count = defaultdict(int)
                for record in unique_detailed_records:
                    pattern = record.get('模式', 'PK10协作')
                    pattern_count[pattern] += 1
                
                main_opposite_type = max(opposite_type_counts.items(), key=lambda x: x[1])[0] if opposite_type_counts else '协作模式'
                
                account_stats_info = []
                for account in account_group:
                    if hasattr(self, 'df_valid') and self.df_valid is not None:
                        account_data = self.df_valid[
                            (self.df_valid['会员账号'] == account) & 
                            (self.df_valid['彩种'] == lottery)
                        ]
                        total_periods = account_data['期号'].nunique()
                        records_count = len(account_data)
                    else:
                        total_periods_stats = self.account_total_periods_by_lottery.get(lottery, {})
                        record_stats = self.account_record_stats_by_lottery.get(lottery, {})
                        total_periods = total_periods_stats.get(account, 0)
                        records_count = record_stats.get(account, 0)
                    
                    account_stats_info.append(f"{account}({total_periods}期/{records_count}记录)")
                
                activity_level = self.get_account_group_activity_level(account_group, lottery)
                
                continuous_pattern = {
                    '账户组': account_group,
                    '彩种': lottery,
                    '彩种类型': records[0]['彩种类型'] if records else 'PK10',
                    '账户数量': len(account_group),
                    '主要对立类型': main_opposite_type,
                    '对立类型分布': dict(opposite_type_counts),
                    '对刷期数': len(unique_detailed_records),
                    '总投注金额': total_investment,
                    '平均相似度': avg_similarity,
                    '模式分布': dict(pattern_count),
                    '详细记录': unique_detailed_records,
                    '账户活跃度': activity_level,
                    '账户统计信息': account_stats_info,
                    '要求最小对刷期数': required_min_periods,
                    '检测类型': records[0].get('检测类型', 'PK10序列位置'),
                    '完整覆盖期数': len(unique_detailed_records),
                    '总检测期数': len(sorted_records)
                }
                
                continuous_patterns.append(continuous_pattern)
        
        return continuous_patterns

    def _detect_single_position_full_coverage(self, period_data, period, specific_lottery='PK10'):
        """增强版：检测单个位置全覆盖模式 - 支持单个位置单独下注和组合位置打包下注"""
        patterns = []
        
        pk10_positions = ['冠军', '亚军', '第三名', '第四名', '第五名', 
                         '第六名', '第七名', '第八名', '第九名', '第十名']
        
        # 按账户收集投注信息
        account_data = defaultdict(lambda: {
            'positions': set(),
            'direction': None,
            'total_amount': 0,
            'position_details': defaultdict(list)
        })
        
        for _, row in period_data.iterrows():
            account = row['会员账号']
            play_category = row.get('玩法分类', '')
            content = row['内容']
            amount = row.get('投注金额', 0)
            direction = row.get('投注方向', '')
            
            if not direction:
                direction = self.enhanced_extract_direction_with_position(content, play_category, 'PK10')
                if not direction:
                    continue
            
            # 确定这个投注覆盖了哪些位置
            positions_covered = []
            
            # 情况1：玩法分类是具体的单个位置
            if play_category in pk10_positions:
                positions_covered.append(play_category)
            
            # 情况2：玩法分类是1-5名或6-10名
            elif play_category == '1-5名':
                positions_covered.extend(['冠军', '亚军', '第三名', '第四名', '第五名'])
            elif play_category == '6-10名':
                positions_covered.extend(['第六名', '第七名', '第八名', '第九名', '第十名'])
            
            # 情况3：从内容中提取位置
            else:
                # 检查内容是否包含位置信息
                content_str = str(content)
                for position in pk10_positions:
                    if position in content_str:
                        positions_covered.append(position)
            
            if not positions_covered:
                continue
            
            # 记录账户的投注信息
            account_info = account_data[account]
            
            # 检查方向是否一致
            if account_info['direction'] is None:
                account_info['direction'] = direction
            elif account_info['direction'] != direction:
                # 方向不一致，跳过这个投注
                continue
            
            # 添加位置和金额
            for position in positions_covered:
                account_info['positions'].add(position)
                account_info['position_details'][position].append({
                    'amount': amount,
                    'content': content,
                    'play_category': play_category
                })
            
            # 累加总金额（注意：这里需要避免重复累加）
            # 对于单个位置单独下注，每个位置都有单独的金额
            # 对于组合位置打包下注，金额应该只加一次
            if len(positions_covered) > 1 and play_category in ['1-5名', '6-10名']:
                # 组合投注，金额只加一次
                account_info['total_amount'] += amount
            else:
                # 单个位置投注，需要计算总金额
                # 这里稍后在汇总时再计算
                pass
        
        # 对于单个位置单独下注，计算总金额
        for account, info in account_data.items():
            if info['direction'] and info['positions']:
                # 如果账户有多个位置投注，但总金额为0，说明是单个位置单独下注
                if info['total_amount'] == 0 and len(info['positions']) > 0:
                    # 计算所有位置的总金额
                    total = 0
                    for position, details in info['position_details'].items():
                        if details:
                            total += details[0]['amount']  # 取第一个投注记录的金额
                    info['total_amount'] = total
        
        # 找出所有账户
        all_accounts = list(account_data.keys())
        if len(all_accounts) < 2:
            return patterns
        
        # 检查任意两个账户的组合
        for i in range(len(all_accounts)):
            for j in range(i+1, len(all_accounts)):
                account1 = all_accounts[i]
                account2 = all_accounts[j]
                
                info1 = account_data[account1]
                info2 = account_data[account2]
                
                # 检查方向是否相同
                if not info1['direction'] or not info2['direction']:
                    continue
                
                if info1['direction'] != info2['direction']:
                    continue
                
                # 检查位置是否互补（没有重叠且合起来覆盖十个位置）
                positions1 = info1['positions']
                positions2 = info2['positions']
                
                if positions1 & positions2:  # 有重叠位置
                    continue
                
                all_covered = positions1 | positions2
                if len(all_covered) != 10:
                    continue
                
                # 检查是否是标准的1-5名和6-10名互补
                is_standard = False
                positions_1_5 = set(['冠军', '亚军', '第三名', '第四名', '第五名'])
                positions_6_10 = set(['第六名', '第七名', '第八名', '第九名', '第十名'])
                
                if (positions1 == positions_1_5 and positions2 == positions_6_10) or \
                   (positions1 == positions_6_10 and positions2 == positions_1_5):
                    is_standard = True
                    account1_positions_desc = '1-5名' if positions1 == positions_1_5 else '6-10名'
                    account2_positions_desc = '6-10名' if positions2 == positions_6_10 else '1-5名'
                else:
                    # 非标准位置分配
                    account1_positions_desc = f"{len(positions1)}个位置"
                    account2_positions_desc = f"{len(positions2)}个位置"
                
                # 获取金额
                amount1 = info1['total_amount']
                amount2 = info2['total_amount']
                
                # 检查金额平衡
                max_ratio = self.config.amount_threshold.get('max_amount_ratio', 10)
                if min(amount1, amount2) == 0:
                    continue
                
                if max(amount1, amount2) / min(amount1, amount2) > max_ratio:
                    continue
                
                # 生成模式描述
                direction_display = info1['direction']
                if direction_display.startswith('数字-'):
                    number = direction_display.replace('数字-', '')
                    pattern_desc = f'PK10十位置协作-数字{number}'
                elif direction_display.startswith('多数字-'):
                    numbers = direction_display.replace('多数字-', '')
                    pattern_desc = f'PK10十位置协作-多数字{numbers}'
                else:
                    pattern_desc = f'PK10十位置协作-{direction_display}'
                
                # 添加标准类型标识
                pattern_type = '标准分组' if is_standard else '非标分组'
                
                record = {
                    '期号': period,
                    '彩种': specific_lottery,
                    '彩种类型': 'PK10',
                    '账户组': [account1, account2],
                    '方向组': [direction_display, direction_display],
                    '玩法分类': [account1_positions_desc, account2_positions_desc],
                    '金额组': [amount1, amount2],
                    '总金额': amount1 + amount2,
                    '相似度': 1.0,
                    '账户数量': 2,
                    '模式': f'PK10十位置{pattern_type}-{direction_display}',
                    '对立类型': f'位置协作-{direction_display}',
                    '检测类型': 'PK10序列位置',
                    '是否互补': True,
                    '位置覆盖详情': {
                        '覆盖类型': '完整覆盖',
                        account1: account1_positions_desc,
                        account2: account2_positions_desc,
                        '详细分配': {
                            account1: sorted(list(positions1)),
                            account2: sorted(list(positions2))
                        }
                    }
                }
                
                patterns.append(record)
        
        return patterns
    
    def _extract_single_position(self, play_category, content):
        """从单个位置投注中提取位置信息"""
        # 首先从玩法分类中提取
        play_str = str(play_category).strip()
        
        if '冠军' in play_str:
            return '冠军'
        elif '亚军' in play_str:
            return '亚军'
        elif '第三名' in play_str or '季军' in play_str or '第3名' in play_str:
            return '第三名'
        elif '第四名' in play_str or '第4名' in play_str:
            return '第四名'
        elif '第五名' in play_str or '第5名' in play_str:
            return '第五名'
        elif '第六名' in play_str or '第6名' in play_str:
            return '第六名'
        elif '第七名' in play_str or '第7名' in play_str:
            return '第七名'
        elif '第八名' in play_str or '第8名' in play_str:
            return '第八名'
        elif '第九名' in play_str or '第9名' in play_str:
            return '第九名'
        elif '第十名' in play_str or '第10名' in play_str:
            return '第十名'
        
        # 如果玩法分类中没有，尝试从内容中提取
        content_str = str(content)
        if '冠军' in content_str:
            return '冠军'
        elif '亚军' in content_str:
            return '亚军'
        elif '第三名' in content_str or '季军' in content_str or '第3名' in content_str:
            return '第三名'
        elif '第四名' in content_str or '第4名' in content_str:
            return '第四名'
        elif '第五名' in content_str or '第5名' in content_str:
            return '第五名'
        elif '第六名' in content_str or '第6名' in content_str:
            return '第六名'
        elif '第七名' in content_str or '第7名' in content_str:
            return '第七名'
        elif '第八名' in content_str or '第8名' in content_str:
            return '第八名'
        elif '第九名' in content_str or '第9名' in content_str:
            return '第九名'
        elif '第十名' in content_str or '第10名' in content_str:
            return '第十名'
        
        return None

    def _detect_arbitrary_position_coverage(self, period_data, period, specific_lottery='PK10'):
        """增强版：检测任意位置分配组合 - 支持单个位置单独下注"""
        patterns = []
        
        pk10_positions = ['冠军', '亚军', '第三名', '第四名', '第五名', 
                         '第六名', '第七名', '第八名', '第九名', '第十名']
        
        # 按账户收集投注信息
        account_data = defaultdict(lambda: {
            'positions': set(),
            'direction': None,
            'total_amount': 0,
            'position_amounts': {}
        })
        
        for _, row in period_data.iterrows():
            account = row['会员账号']
            play_category = row.get('玩法分类', '')
            content = row['内容']
            amount = row.get('投注金额', 0)
            direction = row.get('投注方向', '')
            
            if not direction:
                direction = self.enhanced_extract_direction_with_position(content, play_category, 'PK10')
                if not direction:
                    continue
            
            # 确定这个投注覆盖了哪些位置
            positions_covered = []
            
            # 情况1：玩法分类是具体的单个位置
            if play_category in pk10_positions:
                positions_covered.append(play_category)
            
            # 情况2：玩法分类是1-5名或6-10名
            elif play_category == '1-5名':
                positions_covered.extend(['冠军', '亚军', '第三名', '第四名', '第五名'])
            elif play_category == '6-10名':
                positions_covered.extend(['第六名', '第七名', '第八名', '第九名', '第十名'])
            
            # 情况3：从内容中提取位置
            else:
                content_str = str(content)
                for position in pk10_positions:
                    if position in content_str:
                        positions_covered.append(position)
            
            if not positions_covered:
                continue
            
            # 记录账户的投注信息
            account_info = account_data[account]
            
            # 检查方向是否一致
            if account_info['direction'] is None:
                account_info['direction'] = direction
            elif account_info['direction'] != direction:
                # 方向不一致，跳过这个投注
                continue
            
            # 记录位置和金额
            for position in positions_covered:
                account_info['positions'].add(position)
                if position not in account_info['position_amounts']:
                    account_info['position_amounts'][position] = 0
                account_info['position_amounts'][position] += amount
            
            # 累加总金额
            account_info['total_amount'] += amount
        
        # 找出所有账户
        all_accounts = list(account_data.keys())
        if len(all_accounts) < 2:
            return patterns
        
        # 检查任意两个账户的组合
        for i in range(len(all_accounts)):
            for j in range(i+1, len(all_accounts)):
                account1 = all_accounts[i]
                account2 = all_accounts[j]
                
                info1 = account_data[account1]
                info2 = account_data[account2]
                
                # 检查方向是否相同
                if not info1['direction'] or not info2['direction']:
                    continue
                
                if info1['direction'] != info2['direction']:
                    continue
                
                # 检查位置是否互补（没有重叠且合起来覆盖十个位置）
                positions1 = info1['positions']
                positions2 = info2['positions']
                
                if positions1 & positions2:  # 有重叠位置
                    continue
                
                all_covered = positions1 | positions2
                if len(all_covered) != 10:
                    continue
                
                # 检查金额平衡
                max_ratio = self.config.amount_threshold.get('max_amount_ratio', 10)
                if min(info1['total_amount'], info2['total_amount']) == 0:
                    continue
                
                if max(info1['total_amount'], info2['total_amount']) / min(info1['total_amount'], info2['total_amount']) > max_ratio:
                    continue
                
                # 生成位置描述
                positions_1_5 = set(['冠军', '亚军', '第三名', '第四名', '第五名'])
                positions_6_10 = set(['第六名', '第七名', '第八名', '第九名', '第十名'])
                
                if (positions1 == positions_1_5 and positions2 == positions_6_10) or \
                   (positions2 == positions_1_5 and positions1 == positions_6_10):
                    account1_positions_desc = '1-5名' if positions1 == positions_1_5 else '6-10名'
                    account2_positions_desc = '6-10名' if positions2 == positions_6_10 else '1-5名'
                    pattern_type = '标准分组'
                else:
                    account1_positions_desc = f"{len(positions1)}个位置"
                    account2_positions_desc = f"{len(positions2)}个位置"
                    pattern_type = '非标分组'
                
                # 生成模式描述
                direction_display = info1['direction']
                if direction_display.startswith('数字-'):
                    number = direction_display.replace('数字-', '')
                    pattern_desc = f'PK10十位置{pattern_type}-数字{number}'
                elif direction_display.startswith('多数字-'):
                    numbers = direction_display.replace('多数字-', '')
                    pattern_desc = f'PK10十位置{pattern_type}-多数字{numbers}'
                else:
                    pattern_desc = f'PK10十位置{pattern_type}-{direction_display}'
                
                record = {
                    '期号': period,
                    '彩种': specific_lottery,
                    '彩种类型': 'PK10',
                    '账户组': [account1, account2],
                    '方向组': [direction_display, direction_display],
                    '玩法分类': [account1_positions_desc, account2_positions_desc],
                    '金额组': [info1['total_amount'], info2['total_amount']],
                    '总金额': info1['total_amount'] + info2['total_amount'],
                    '相似度': 1.0,
                    '账户数量': 2,
                    '模式': pattern_desc,
                    '对立类型': f'位置协作-{direction_display}',
                    '检测类型': 'PK10序列位置',
                    '是否互补': True,
                    '位置覆盖详情': {
                        '覆盖类型': '完整覆盖',
                        account1: account1_positions_desc,
                        account2: account2_positions_desc,
                        '详细分配': {
                            account1: sorted(list(positions1)),
                            account2: sorted(list(positions2))
                        }
                    }
                }
                
                patterns.append(record)
        
        return patterns

    def _check_individual_position_coverage(self, account_position_bets, account1, account2, period):
        """检查两个账户的单个位置注单协作"""
        result = {
            'covered': False,
            'patterns': []
        }
        
        position_coverage = {}
        common_directions = set()
        
        for position in self.pk10_positions:
            account1_bets = account_position_bets[account1].get(position, [])
            account2_bets = account_position_bets[account2].get(position, [])
            
            if not account1_bets or not account2_bets:
                continue
            
            account1_content = account1_bets[0]['content']
            account2_content = account2_bets[0]['content']
            
            if account1_content == account2_content:
                position_coverage[position] = {
                    'content': account1_content,
                    'account1_amount': sum(bet['amount'] for bet in account1_bets),
                    'account2_amount': sum(bet['amount'] for bet in account2_bets)
                }
                common_directions.add(account1_content)
        
        if len(position_coverage) == len(self.pk10_positions) and len(common_directions) == 1:
            common_direction = list(common_directions)[0]
            total_amount = sum(pos_info['account1_amount'] + pos_info['account2_amount'] 
                              for pos_info in position_coverage.values())
            
            account1_positions = []
            account2_positions = []
            for position in self.pk10_positions:
                if account_position_bets[account1].get(position):
                    account1_positions.append(position)
                if account_position_bets[account2].get(position):
                    account2_positions.append(position)
            
            pattern = {
                '期号': period,
                '彩种': 'PK10',
                '彩种类型': 'PK10',
                '账户组': [account1, account2],
                '方向组': [common_direction, common_direction],
                '金额组': [
                    sum(account_position_bets[account1][pos][0]['amount'] for pos in account1_positions),
                    sum(account_position_bets[account2][pos][0]['amount'] for pos in account2_positions)
                ],
                '总金额': total_amount,
                '相似度': 1.0,
                '账户数量': 2,
                '模式': f'PK10十位置全覆盖-{common_direction}',
                '对立类型': f'全覆盖协作-{common_direction}',
                '检测类型': 'PK10序列位置',
                '位置分配': {
                    account1: account1_positions,
                    account2: account2_positions
                }
            }
            
            result['covered'] = True
            result['patterns'].append(pattern)
        
        return result
    
    def _extract_position_from_play_category(self, play_category):
        """从玩法分类中提取位置信息 - 增强版"""
        play_str = str(play_category).strip()
        
        position_mapping = {
            # 冠军
            '冠军': '冠军',
            '第1名': '冠军',
            '第一名': '冠军',
            '前一': '冠军',
            '冠 军': '冠军',
            '冠　军': '冠军',
            
            # 亚军
            '亚军': '亚军',
            '第2名': '亚军',
            '第二名': '亚军',
            '前二': '亚军',
            '亚 军': '亚军',
            '亚　军': '亚军',
            
            # 第三名
            '季军': '第三名',
            '第3名': '第三名',
            '第三名': '第三名',
            '前三': '第三名',
            
            # 第四名
            '第4名': '第四名',
            '第四名': '第四名',
            '前四': '第四名',
            
            # 第五名
            '第5名': '第五名',
            '第五名': '第五名',
            '前五': '第五名',
            
            # 第六名
            '第6名': '第六名',
            '第六名': '第六名',
            
            # 第七名
            '第7名': '第七名',
            '第七名': '第七名',
            
            # 第八名
            '第8名': '第八名',
            '第八名': '第八名',
            
            # 第九名
            '第9名': '第九名',
            '第九名': '第九名',
            
            # 第十名
            '第10名': '第十名',
            '第十名': '第十名'
        }
        
        for key, position in position_mapping.items():
            if key in play_str:
                return position
        
        return ''
    
    def _detect_1_5_6_10_collaboration(self, period_data, period, specific_lottery='PK10'):
        """修复版：检测1-5名和6-10名的协作模式 - 添加位置信息"""
        patterns = []
        
        play_1_5 = period_data[period_data['玩法分类'] == '1-5名']
        play_6_10 = period_data[period_data['玩法分类'] == '6-10名']
        
        if len(play_1_5) == 0 or len(play_6_10) == 0:
            return patterns
        
        # 按账户分组
        account_1_5_data = {}
        account_6_10_data = {}
        
        # 处理1-5名数据
        for _, row in play_1_5.iterrows():
            account = row['会员账号']
            direction = row.get('投注方向', '')
            amount = row.get('投注金额', 0)
            content = row['内容']
            
            if direction:
                account_1_5_data[account] = {
                    'direction': direction,
                    'amount': amount,
                    'content': content,
                    'play_category': '1-5名'
                }
        
        # 处理6-10名数据
        for _, row in play_6_10.iterrows():
            account = row['会员账号']
            direction = row.get('投注方向', '')
            amount = row.get('投注金额', 0)
            content = row['内容']
            
            if direction:
                account_6_10_data[account] = {
                    'direction': direction,
                    'amount': amount,
                    'content': content,
                    'play_category': '6-10名'
                }
        
        # 查找协作模式
        for acc1, data1 in account_1_5_data.items():
            for acc2, data2 in account_6_10_data.items():
                if acc1 == acc2:
                    continue
                
                # 检查投注方向是否相同
                if data1['direction'] != data2['direction']:
                    continue
                
                # 检查金额平衡
                max_ratio = self.config.amount_threshold.get('max_amount_ratio', 10)
                if max(data1['amount'], data2['amount']) / min(data1['amount'], data2['amount']) > max_ratio:
                    continue
                
                account_group = [acc1, acc2]
                directions = [data1['direction'], data2['direction']]
                amounts = [data1['amount'], data2['amount']]
                total_amount = data1['amount'] + data2['amount']
                
                # 提取投注内容
                if data1['direction'].startswith('数字-'):
                    numbers = data1['direction'].replace('数字-', '')
                    pattern_desc = f'PK10十位置协作-数字{numbers}'
                elif data1['direction'].startswith('多数字-'):
                    numbers = data1['direction'].replace('多数字-', '')
                    pattern_desc = f'PK10十位置协作-多数字{numbers}'
                else:
                    pattern_desc = f'PK10十位置协作-{data1["direction"]}'
                
                record = {
                    '期号': period,
                    '彩种': specific_lottery,
                    '彩种类型': 'PK10',
                    '账户组': account_group,
                    '方向组': directions,
                    '玩法分类': ['1-5名', '6-10名'],  # 添加位置信息
                    '金额组': amounts,
                    '总金额': total_amount,
                    '相似度': 1.0,
                    '账户数量': 2,
                    '模式': pattern_desc,
                    '对立类型': f'位置协作-{data1["direction"]}',
                    '检测类型': 'PK10序列位置'
                }
                
                patterns.append(record)
        
        return patterns

    def _detect_vertical_format_collaboration(self, period_data, period, specific_lottery='PK10'):
        """修复版：检测竖线分隔格式的协作模式 - 确保添加位置信息"""
        patterns = []
        
        # 查找使用竖线分隔的内容
        vertical_bets = period_data[period_data['内容'].str.contains('|', na=False, regex=False)]
        
        if len(vertical_bets) < 2:
            return patterns
        
        # 按账户分组
        account_bets = {}
        for _, row in vertical_bets.iterrows():
            account = row['会员账号']
            content = row['内容']
            direction = row.get('投注方向', '')
            amount = row.get('投注金额', 0)
            play_category = row.get('玩法分类', '')
            original_play = row.get('玩法', '')  # 原始玩法字段
            
            if account not in account_bets:
                account_bets[account] = []
            
            account_bets[account].append({
                'content': content,
                'direction': direction,
                'amount': amount,
                'play_category': play_category,
                'original_play': original_play
            })
        
        # 比较账户间的投注内容
        accounts = list(account_bets.keys())
        if len(accounts) < 2:
            return patterns
        
        # 去重集合，避免同一期号重复检测相同账户对
        detected_pairs = set()
        
        for i in range(len(accounts)):
            for j in range(i+1, len(accounts)):
                acc1 = accounts[i]
                acc2 = accounts[j]
                
                bets1 = account_bets[acc1]
                bets2 = account_bets[acc2]
                
                # 检查是否有相同方向的对刷
                for bet1 in bets1:
                    for bet2 in bets2:
                        if bet1['direction'] and bet2['direction'] and bet1['direction'] == bet2['direction']:
                            # 创建去重键
                            pair_key = (period, tuple(sorted([acc1, acc2])), bet1['direction'])
                            if pair_key in detected_pairs:
                                continue
                            
                            detected_pairs.add(pair_key)
                            
                            # 检查金额平衡
                            max_ratio = self.config.amount_threshold.get('max_amount_ratio', 10)
                            if max(bet1['amount'], bet2['amount']) / min(bet1['amount'], bet2['amount']) > max_ratio:
                                continue
                            
                            # 获取玩法分类
                            play1 = bet1['original_play'] or bet1['play_category']
                            play2 = bet2['original_play'] or bet2['play_category']
                            
                            # 判断是否互补
                            is_complementary = False
                            
                            # 检查play1和play2是否一个包含1-5，另一个包含6-10
                            play1_str = str(play1).lower()
                            play2_str = str(play2).lower()
                            
                            # 定义1-5名的关键词
                            one_to_five_keywords = ['1-5名', '第1~5名', '定位胆_第1~5名', '冠军', '亚军', '第三名', '第四名', '第五名', '第1名', '第2名', '第3名', '第4名', '第5名']
                            # 定义6-10名的关键词  
                            six_to_ten_keywords = ['6-10名', '第6~10名', '定位胆_第6~10名', '第六名', '第七名', '第八名', '第九名', '第十名', '第6名', '第7名', '第8名', '第9名', '第10名']
                            
                            # 检查play1是否包含1-5关键词，play2是否包含6-10关键词
                            condition1 = any(keyword in play1_str for keyword in one_to_five_keywords) and any(keyword in play2_str for keyword in six_to_ten_keywords)
                            # 检查play1是否包含6-10关键词，play2是否包含1-5关键词
                            condition2 = any(keyword in play1_str for keyword in six_to_ten_keywords) and any(keyword in play2_str for keyword in one_to_five_keywords)
                            
                            is_complementary = condition1 or condition2
                            
                            # 关键修复：只保留互补的投注
                            if not is_complementary:
                                continue  # 跳过不互补的投注
                            
                            # 确定模式类型
                            pattern_type = 'PK10完整协作'
                            detection_type = 'PK10序列位置'
                            
                            # 获取投注位置详情
                            position_detail1 = self._get_position_detail(play1, play1)
                            position_detail2 = self._get_position_detail(play2, play2)
                            
                            record = {
                                '期号': period,
                                '彩种': specific_lottery,
                                '彩种类型': 'PK10',
                                '账户组': [acc1, acc2],
                                '方向组': [bet1['direction'], bet2['direction']],
                                '玩法分类': [position_detail1, position_detail2],  # 使用简化的位置详情
                                '金额组': [bet1['amount'], bet2['amount']],
                                '总金额': bet1['amount'] + bet2['amount'],
                                '相似度': 1.0,
                                '账户数量': 2,
                                '模式': f'{pattern_type}-{bet1["direction"]}',
                                '对立类型': f'{pattern_type.replace("PK10", "")}-{bet1["direction"]}',
                                '检测类型': detection_type,
                                '是否互补': is_complementary
                            }
                            
                            # 判断哪个账户投1-5名，哪个投6-10名
                            if any(keyword in play1_str for keyword in one_to_five_keywords):
                                acc1_positions = '1-5名'
                                acc2_positions = '6-10名'
                            else:
                                acc1_positions = '6-10名'
                                acc2_positions = '1-5名'
                            
                            record['位置覆盖详情'] = {
                                '覆盖类型': '完整覆盖',
                                acc1: acc1_positions,
                                acc2: acc2_positions
                            }
                            
                            patterns.append(record)
        
        return patterns
    
    def _get_position_detail(self, play_category, original_play):
        """获取位置详情 - 修正版"""
        # 首先检查原始玩法
        original_str = str(original_play).lower() if original_play else ""
        play_str = str(play_category).lower() if play_category else ""
        
        # 检查是否是"定位胆_第1~5名"格式
        if '定位胆_第1~5名' in original_str or '定位胆_第1~5名' in play_str:
            return '1-5名'
        elif '定位胆_第6~10名' in original_str or '定位胆_第6~10名' in play_str:
            return '6-10名'
        # 检查是否是"定位胆"简写
        elif '定位胆' in original_str or '定位胆' in play_str:
            # 尝试从内容推断具体位置
            return '定位胆'  # 暂时返回通用名称
        elif '1-5名' in original_str or '1-5名' in play_str:
            return '1-5名'
        elif '6-10名' in original_str or '6-10名' in play_str:
            return '6-10名'
        # 检查具体位置
        elif any(pos in original_str for pos in ['冠军', '第1名', '第一名', '前一']):
            return '冠军'
        elif any(pos in original_str for pos in ['亚军', '第2名', '第二名', '前二']):
            return '亚军'
        elif any(pos in original_str for pos in ['第三名', '第3名', '季军', '前三']):
            return '第三名'
        elif any(pos in original_str for pos in ['第四名', '第4名', '前四']):
            return '第四名'
        elif any(pos in original_str for pos in ['第五名', '第5名', '前五']):
            return '第五名'
        elif any(pos in original_str for pos in ['第六名', '第6名']):
            return '第六名'
        elif any(pos in original_str for pos in ['第七名', '第7名']):
            return '第七名'
        elif any(pos in original_str for pos in ['第八名', '第8名']):
            return '第八名'
        elif any(pos in original_str for pos in ['第九名', '第9名']):
            return '第九名'
        elif any(pos in original_str for pos in ['第十名', '第10名']):
            return '第十名'
        else:
            # 如果还是无法确定，返回玩法分类
            return play_category if play_category else original_play

    def find_continuous_sequence_patterns(self, sequence_patterns):
        """查找连续的序列模式"""
        if not sequence_patterns:
            return []
        
        account_group_patterns = defaultdict(list)
        for pattern in sequence_patterns:
            key = (tuple(sorted(pattern['账户组'])), pattern['投注内容'])
            account_group_patterns[key].append(pattern)
        
        continuous_patterns = []
        
        for (account_group, bet_content), records in account_group_patterns.items():
            sorted_records = sorted(records, key=lambda x: x['期号'])
            
            if len(sorted_records) >= 3:
                total_investment = sum(r['总投注金额'] for r in sorted_records)
                coverage_ratios = [r['覆盖度'] for r in sorted_records]
                avg_coverage = np.mean(coverage_ratios) if coverage_ratios else 0
                
                if avg_coverage >= 1.0:
                    account_count = len(account_group)
                    if account_count in [2, 3]:
                        continuous_patterns.append({
                            '账户组': list(account_group),
                            '账户数量': account_count,
                            '投注内容': bet_content,
                            '彩种': 'PK10',
                            '彩种类型': 'PK10',
                            '连续期数': len(sorted_records),
                            '总投注金额': total_investment,
                            '平均覆盖度': avg_coverage,
                            '详细记录': sorted_records,
                            '模式类型': '序列覆盖',
                            '模式描述': f'PK10十位置全覆盖-{bet_content}({account_count}账户协作)',
                            '检测类型': 'PK10序列位置'
                        })
        
        return continuous_patterns

    def display_pk10_sequence_results(self, patterns):
        """显示PK10序列检测结果"""
        if not patterns:
            return
        
        st.subheader("🎯 PK10序列位置检测结果")
        
        total_groups = len(patterns)
        total_periods = sum(p['连续期数'] for p in patterns)
        total_amount = sum(p['总投注金额'] for p in patterns)
        
        account_count_stats = defaultdict(int)
        for pattern in patterns:
            account_count_stats[pattern['账户数量']] += 1
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("序列对刷组数", total_groups)
        with col2:
            st.metric("总对刷期数", total_periods)
        with col3:
            st.metric("总涉及金额", f"¥{total_amount:,.2f}")
        with col4:
            two_account = account_count_stats.get(2, 0)
            three_account = account_count_stats.get(3, 0)
            st.metric("账户组合", f"2账户:{two_account}组 3账户:{three_account}组")
        
        content_stats = defaultdict(int)
        for pattern in patterns:
            content_stats[pattern['投注内容']] += 1
        
        st.write("**投注内容分布:**")
        content_cols = st.columns(min(5, len(content_stats)))
        for i, (content, count) in enumerate(content_stats.items()):
            if i < len(content_cols):
                with content_cols[i]:
                    st.metric(f"{content}模式", f"{count}组")
        
        st.info("""
        **检测模式说明：**
        - **2账户协作**：两个账户共同覆盖PK10十个位置，投注相同内容
        - **3账户协作**：三个账户共同覆盖PK10十个位置，投注相同内容
        - **十个位置全覆盖**：确保PK10的十个位置都被相同内容覆盖
        - **连续多期出现**：要求至少连续3期出现相同模式
        """)
        
        for i, pattern in enumerate(patterns, 1):
            st.markdown(f"**对刷组 {i}:** {' ↔ '.join(pattern['账户组'])}")
            account_type = "2账户协作" if pattern['账户数量'] == 2 else "3账户协作"
            st.markdown(f"**模式类型:** {account_type} | **投注内容:** {pattern['投注内容']} | **连续期数:** {pattern['连续期数']}期")
            st.markdown(f"**总金额:** ¥{pattern['总投注金额']:,.2f} | **平均覆盖度:** {pattern['平均覆盖度']:.1%}")
            
            st.markdown("**详细记录:**")
            for j, record in enumerate(pattern['详细记录'], 1):
                position_coverage = []
                for pos_record in record['位置详情']:
                    position_coverage.append(f"{pos_record['position']}({','.join(pos_record['accounts'])})")
                
                st.write(f"{j}. 期号: {record['期号']} | 覆盖位置: {record['覆盖位置数']}/{record['总位置数']} | 金额: ¥{record['总投注金额']:,.2f}")
                st.write(f"   位置分配: {' | '.join(position_coverage)}")
            
            if i < len(patterns):
                st.markdown("---")

    def _calculate_detailed_account_stats(self, patterns):
        """彻底修复的账户统计计算方法 - 去掉涉及彩种列"""
        account_participation = defaultdict(lambda: {
            'groups': set(),
            'lotteries': set(),
            'wash_periods': set(),
            'lottery_wash_periods': defaultdict(set),  # 按彩种记录对刷期数
            'total_bet_amount': 0,
        })
        
        if not hasattr(self, 'df_valid') or self.df_valid is None:
            return []
        
        # 收集账户参与信息
        for pattern in patterns:
            group_id = f"组{len(account_participation) + 1}"
            lottery = pattern['彩种']
            
            for account in pattern['账户组']:
                account_info = account_participation[account]
                account_info['groups'].add(group_id)
                account_info['lotteries'].add(lottery)
                
                # 记录每个彩种的对刷期数
                for record in pattern['详细记录']:
                    account_info['wash_periods'].add(record['期号'])
                    account_info['lottery_wash_periods'][lottery].add(record['期号'])
                
                # 计算该账户在这个模式中的总投注金额
                pattern_bet_amount = 0
                for record in pattern['详细记录']:
                    for acc, amt in zip(record['账户组'], record['金额组']):
                        if acc == account:
                            pattern_bet_amount += amt
                
                account_info['total_bet_amount'] += pattern_bet_amount
        
        # 生成统计记录
        account_stats = []
        for account, info in account_participation.items():
            groups_count = len(info['groups'])
            wash_periods_count = len(info['wash_periods'])
            total_bet_amount = info['total_bet_amount']
            avg_period_amount = total_bet_amount / wash_periods_count if wash_periods_count > 0 else 0
            
            # 计算彩种总投注期数
            lottery_total_periods = 0
            
            for detected_lottery in info['lotteries']:
                account_all_data = self.df_valid[self.df_valid['会员账号'] == account]
                
                if '原始彩种' in self.df_valid.columns:
                    account_lottery_data = account_all_data[account_all_data['原始彩种'] == detected_lottery]
                else:
                    account_lottery_data = account_all_data[account_all_data['彩种'] == detected_lottery]
                
                if len(account_lottery_data) == 0 and '彩种类型' in self.df_valid.columns:
                    account_lottery_data = account_all_data[account_all_data['彩种类型'] == detected_lottery]
                
                if len(account_lottery_data) == 0:
                    account_lottery_data = account_all_data[account_all_data['彩种'].str.contains(detected_lottery, na=False)]
                
                lottery_total_periods += account_lottery_data['期号'].nunique()
            
            # 生成违规彩种字符串（彩种（期数））
            violation_lotteries = []
            for lottery, periods in info['lottery_wash_periods'].items():
                period_count = len(periods)
                violation_lotteries.append(f"{lottery}（{period_count}期）")
            
            violation_lotteries_str = "；".join(violation_lotteries)
            
            stat_record = {
                '账户': account,
                '参与组合数': groups_count,
                '彩种总投注期数': lottery_total_periods,
                '实际对刷期数': wash_periods_count,
                '违规彩种（彩种（期数））': violation_lotteries_str,
                '总投注金额': total_bet_amount,
                '平均每期金额': avg_period_amount
            }
            
            account_stats.append(stat_record)
        
        return sorted(account_stats, key=lambda x: x['参与组合数'], reverse=True)

    def exclude_multi_direction_accounts(self, df_valid):
        """排除同一账户多方向下注"""
        if '玩法分类' not in df_valid.columns:
            return df_valid
        
        pk10_positions = ['冠军', '亚军', '第三名', '第四名', '第五名', 
                         '第六名', '第七名', '第八名', '第九名', '第十名']
        
        single_position_mask = df_valid['玩法分类'].isin(pk10_positions)
        
        single_position_data = df_valid[single_position_mask]
        other_data = df_valid[~single_position_mask]
        
        if len(other_data) > 0:
            if '投注方向' in other_data.columns:
                multi_direction_mask = (
                    other_data.groupby(['期号', '会员账号'])['投注方向']
                    .transform('nunique') > 1
                )
                other_data_filtered = other_data[~multi_direction_mask]
            else:
                other_data_filtered = other_data
        else:
            other_data_filtered = other_data
        
        df_filtered = pd.concat([single_position_data, other_data_filtered], ignore_index=True)
        
        return df_filtered
    
    def get_account_group_activity_level(self, account_group, lottery):
        """获取活跃度水平"""
        if hasattr(self, 'df_valid') and self.df_valid is not None:
            min_total_periods = float('inf')
            
            for account in account_group:
                account_data = self.df_valid[
                    (self.df_valid['会员账号'] == account) & 
                    (self.df_valid['彩种'] == lottery)
                ]
                periods = account_data['期号'].nunique()
                if periods < min_total_periods:
                    min_total_periods = periods
            
            if min_total_periods != float('inf'):
                return self._calculate_activity_level(min_total_periods)
        
        if lottery in self.account_total_periods_by_lottery:
            total_periods_stats = self.account_total_periods_by_lottery[lottery]
            
            account_periods = [total_periods_stats.get(account, 0) for account in account_group]
            if account_periods:
                min_total_periods = min(account_periods)
                return self._calculate_activity_level(min_total_periods)
        
        return 'unknown'
    
    def _calculate_activity_level(self, min_total_periods):
        """根据期数计算活跃度水平"""
        if min_total_periods <= self.config.period_thresholds['low_activity']:
            return 'low'
        elif min_total_periods <= self.config.period_thresholds['medium_activity_high']:
            return 'medium'
        elif min_total_periods <= self.config.period_thresholds['high_activity_low']:
            return 'high'
        else:
            return 'very_high'
    
    def get_required_min_periods(self, account_group, lottery):
        """根据新的活跃度阈值获取所需的最小对刷期数"""
        activity_level = self.get_account_group_activity_level(account_group, lottery)
        
        if activity_level == 'low':
            return self.config.period_thresholds['min_periods_low']
        elif activity_level == 'medium':
            return self.config.period_thresholds['min_periods_medium']
        elif activity_level == 'high':
            return self.config.period_thresholds['min_periods_high']
        else:
            return self.config.period_thresholds['min_periods_very_high']
    
    def display_performance_stats(self):
        """显示性能统计"""
        if not self.performance_stats:
            return
        
        with st.expander("📈 性能统计", expanded=False):
            st.write(f"**数据处理统计:**")
            st.write(f"- 总记录数: {self.performance_stats['total_records']:,}")
            st.write(f"- 总期号数: {self.performance_stats['total_periods']:,}")
            st.write(f"- 总账户数: {self.performance_stats['total_accounts']:,}")
            
            if 'detection_time' in self.performance_stats:
                st.write(f"**检测性能:**")
                st.write(f"- 检测时间: {self.performance_stats['detection_time']:.2f} 秒")
                st.write(f"- 发现模式: {self.performance_stats['total_patterns']} 个")

    def enhanced_analyze_opposite_patterns(self, patterns):
        """增强对立模式分析"""
        if not patterns:
            return {}
        
        analysis = {
            'total_groups': len(patterns),
            'opposite_type_stats': defaultdict(int),
            'position_stats': defaultdict(int),
            'combination_type_stats': defaultdict(int),
            'lottery_opposite_stats': defaultdict(lambda: defaultdict(int))
        }
        
        for pattern in patterns:
            main_opposite = pattern['主要对立类型']
            analysis['opposite_type_stats'][main_opposite] += 1
            
            for record in pattern['详细记录']:
                for direction in record['方向组']:
                    if '-' in direction:
                        position = direction.split('-')[0]
                        analysis['position_stats'][position] += 1
            
            lottery = pattern['彩种']
            analysis['lottery_opposite_stats'][lottery][main_opposite] += 1
        
        return analysis
    
    def display_enhanced_opposite_analysis(self, patterns):
        """显示增强的对立模式分析"""
        if not patterns:
            return
        
        analysis = self.enhanced_analyze_opposite_patterns(patterns)
        
        st.subheader("🎯 对立模式深度分析")
        
        with st.expander("📊 对立类型分布", expanded=True):
            col1, col2 = st.columns(2)
            
            with col1:
                st.write("**主要对立类型:**")
                for opposite_type, count in sorted(analysis['opposite_type_stats'].items(), key=lambda x: x[1], reverse=True)[:10]:
                    st.write(f"- {opposite_type}: {count}组")
            
            with col2:
                st.write("**位置分布:**")
                for position, count in sorted(analysis['position_stats'].items(), key=lambda x: x[1], reverse=True)[:10]:
                    st.write(f"- {position}: {count}次")
        
        with st.expander("🎲 彩种对立模式分析", expanded=False):
            for lottery, opposite_stats in analysis['lottery_opposite_stats'].items():
                st.write(f"**{lottery}:**")
                for opposite_type, count in sorted(opposite_stats.items(), key=lambda x: x[1], reverse=True)[:5]:
                    st.write(f"  - {opposite_type}: {count}组")

    def diagnose_account_data(self, account, lottery):
        """诊断账户数据匹配问题"""
        if not hasattr(self, 'df_valid') or self.df_valid is None:
            return "无数据"
        
        account_data = self.df_valid[self.df_valid['会员账号'] == account]
        
        result = {
            'account': account,
            'total_records': len(account_data),
            'total_periods': account_data['期号'].nunique(),
            'available_columns': self.df_valid.columns.tolist(),
            'lottery_match_attempts': {}
        }
        
        if '原始彩种' in self.df_valid.columns:
            original_match = account_data[account_data['原始彩种'] == lottery]
            result['lottery_match_attempts']['原始彩种精确匹配'] = {
                'records': len(original_match),
                'periods': original_match['期号'].nunique()
            }
        
        if '彩种' in self.df_valid.columns:
            lottery_match = account_data[account_data['彩种'] == lottery]
            result['lottery_match_attempts']['彩种精确匹配'] = {
                'records': len(lottery_match),
                'periods': lottery_match['期号'].nunique()
            }
        
        if '彩种类型' in self.df_valid.columns:
            type_match = account_data[account_data['彩种类型'] == lottery]
            result['lottery_match_attempts']['彩种类型匹配'] = {
                'records': len(type_match),
                'periods': type_match['期号'].nunique()
            }
        
        if '彩种' in account_data.columns:
            result['all_lotteries'] = account_data['彩种'].unique().tolist()
        
        return result

    def display_detailed_results(self, patterns):
        """显示详细检测结果 - 修复彩种统计"""
        if not patterns:
            st.error("❌ 未发现符合阈值条件的连续对刷模式")
            return
        
        # ========== 总体统计 ==========
        st.subheader("📊 总体统计")
        
        # 修复彩种统计逻辑
        lottery_stats = defaultdict(int)
        for pattern in patterns:
            # 使用pattern中的'彩种'字段
            lottery = pattern.get('彩种', '未知')
            lottery_stats[lottery] += 1
        
        # 计算基础统计
        total_groups = len(patterns)
        total_accounts = sum(p['账户数量'] for p in patterns)
        total_wash_periods = sum(p['对刷期数'] for p in patterns)
        total_amount = sum(p['总投注金额'] for p in patterns)
        
        # 第一行：基础数据统计
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("总对刷组数", f"{total_groups}")
        
        with col2:
            st.metric("涉及账户数", f"{total_accounts}")
        
        with col3:
            st.metric("总对刷期数", f"{total_wash_periods}")
        
        with col4:
            st.metric("总涉及金额", f"¥{total_amount:,.2f}")
        
        # ========== 彩种类型统计 ==========
        st.subheader("🎲 彩种类型统计")
        
        # 按数量排序
        sorted_lotteries = sorted(lottery_stats.items(), key=lambda x: x[1], reverse=True)
        
        if sorted_lotteries:
            # 显示所有彩种，不分折叠
            st.write("**各彩种对刷组数:**")
            
            # 计算每行显示多少个
            max_per_row = 6
            num_lotteries = len(sorted_lotteries)
            
            # 如果彩种数量少，显示在一行
            if num_lotteries <= max_per_row:
                cols = st.columns(num_lotteries)
                for i, (lottery, count) in enumerate(sorted_lotteries):
                    with cols[i]:
                        display_name = lottery
                        if len(display_name) > 10:
                            display_name = display_name[:8] + "..."
                        
                        st.metric(
                            label=display_name,
                            value=f"{count}组",
                            help=f"完整名称: {lottery}"
                        )
            else:
                # 如果彩种数量多，分多行显示
                num_rows = (num_lotteries + max_per_row - 1) // max_per_row
                
                for row in range(num_rows):
                    start_idx = row * max_per_row
                    end_idx = min((row + 1) * max_per_row, num_lotteries)
                    row_lotteries = sorted_lotteries[start_idx:end_idx]
                    
                    # 创建这一行的列
                    cols = st.columns(len(row_lotteries))
                    
                    for i, (lottery, count) in enumerate(row_lotteries):
                        with cols[i]:
                            display_name = lottery
                            if len(display_name) > 10:
                                display_name = display_name[:8] + "..."
                            
                            st.metric(
                                label=display_name,
                                value=f"{count}组",
                                help=f"完整名称: {lottery}"
                            )
        
        # ========== 参与账户详细统计 ==========
        st.subheader("👥 参与账户详细统计")
        
        account_stats = self._calculate_detailed_account_stats(patterns)
        
        if account_stats:
            df_stats = pd.DataFrame(account_stats)
            
            # 格式化金额列
            df_stats['总投注金额'] = df_stats['总投注金额'].apply(lambda x: f"¥{x:,.2f}")
            df_stats['平均每期金额'] = df_stats['平均每期金额'].apply(lambda x: f"¥{x:,.2f}")
            
            # 新的列顺序（去掉涉及彩种）
            desired_columns = ['账户', '参与组合数', '彩种总投注期数', '实际对刷期数', 
                              '违规彩种（彩种（期数））', '总投注金额', '平均每期金额']
            
            # 只保留存在的列
            available_columns = [col for col in desired_columns if col in df_stats.columns]
            df_stats = df_stats[available_columns]
            
            st.dataframe(
                df_stats,
                use_container_width=True,
                hide_index=True,
                height=min(400, len(df_stats) * 35 + 38)
            )
        
        # ========== 详细对刷组分析 ==========
        st.subheader("🔍 详细对刷组分析")
        
        patterns_by_lottery = defaultdict(list)
        for pattern in patterns:
            lottery = pattern['彩种']
            patterns_by_lottery[lottery].append(pattern)
        
        for lottery, lottery_patterns in patterns_by_lottery.items():
            total_groups_in_lottery = len(lottery_patterns)
            
            lottery_icon = "🎲"
            if '快三' in lottery or 'K3' in lottery:
                lottery_icon = "🎲"
            elif '六合彩' in lottery or 'LHC' in lottery:
                lottery_icon = "🎰"
            elif 'PK10' in lottery or '赛车' in lottery:
                lottery_icon = "🏁"
            elif '时时彩' in lottery or 'SSC' in lottery:
                lottery_icon = "⏰"
            elif '3D' in lottery or '排列' in lottery:
                lottery_icon = "🔢"
            
            with st.expander(f"{lottery_icon} 彩种：{lottery}（发现{total_groups_in_lottery}组）", expanded=True):
                for i, pattern in enumerate(lottery_patterns, 1):
                    self._display_single_pattern_by_lottery(pattern, i, lottery)
    
    def _display_single_pattern_by_lottery(self, pattern, index, lottery):
        """按彩种显示单个对刷组详情 - 显示所有模式"""
        # 不再过滤任何模式，显示所有检测到的对刷组
        st.markdown(f"**对刷组 {index}:** {' ↔ '.join(pattern['账户组'])}")
        
        activity_icon = "🟢" if pattern['账户活跃度'] == 'low' else "🟡" if pattern['账户活跃度'] == 'medium' else "🟠" if pattern['账户活跃度'] == 'high' else "🔴"
        activity_text = {
            'low': '低活跃度', 
            'medium': '中活跃度', 
            'high': '高活跃度', 
            'very_high': '极高活跃度'
        }.get(pattern['账户活跃度'], pattern['账户活跃度'])
        
        main_type = pattern['主要对立类型']
        if ' vs ' in main_type:
            main_type_parts = main_type.split(' vs ')
            if len(main_type_parts) == 2:
                dir1 = main_type_parts[0].split('(')[0] if '(' in main_type_parts[0] else main_type_parts[0]
                dir2 = main_type_parts[1].split('(')[0] if '(' in main_type_parts[1] else main_type_parts[1]
                display_type = f"{dir1}-{dir2}"
            else:
                display_type = main_type.split('(')[0] if '(' in main_type else main_type
        else:
            display_type = main_type.split('(')[0] if '(' in main_type else main_type
        
        st.markdown(f"**活跃度:** {activity_icon} {activity_text} | **彩种:** {lottery} | **主要类型:** {display_type}")
        
        account_stats_info = []
        for account in pattern['账户组']:
            if hasattr(self, 'df_valid') and self.df_valid is not None:
                account_all_data = self.df_valid[self.df_valid['会员账号'] == account]
                
                account_lottery_data = pd.DataFrame()
                
                if '原始彩种' in self.df_valid.columns:
                    account_lottery_data = account_all_data[account_all_data['原始彩种'] == lottery]
                
                if len(account_lottery_data) == 0 and '彩种' in self.df_valid.columns:
                    account_lottery_data = account_all_data[account_all_data['彩种'] == lottery]
                
                if len(account_lottery_data) == 0 and '彩种类型' in self.df_valid.columns:
                    account_lottery_data = account_all_data[account_all_data['彩种类型'] == lottery]
                
                if len(account_lottery_data) == 0:
                    account_lottery_data = account_all_data[account_all_data['彩种'].str.contains(lottery, na=False)]
                
                total_periods = account_lottery_data['期号'].nunique()
                records_count = len(account_lottery_data)
                
                if total_periods == 0:
                    all_lotteries = account_all_data['彩种'].unique() if '彩种' in account_all_data.columns else []
                    account_stats_info.append(f"{account}(0期/0记录) [实际彩种: {', '.join(all_lotteries[:3])}]")
                else:
                    account_stats_info.append(f"{account}({total_periods}期/{records_count}记录)")
            else:
                account_stats_info.append(f"{account}(数据不可用)")
        
        st.markdown(f"**账户在该彩种投注期数/总记录数:** {', '.join(account_stats_info)}")
        
        st.markdown(f"**对刷期数:** {pattern['对刷期数']}期 (要求≥{pattern['要求最小对刷期数']}期)")
        
        detect_type = pattern.get('检测类型', '传统对刷')
        if detect_type == 'PK10序列位置':
            st.markdown(f"**总金额:** {pattern['总投注金额']:.2f}元")
        else:
            st.markdown(f"**总金额:** {pattern['总投注金额']:.2f}元 | **平均匹配:** {pattern['平均相似度']:.2%}")
        
        st.markdown("**详细记录:**")
        
        # 确保详细记录不重复
        seen_periods = set()
        record_count = 0
        
        for record in pattern['详细记录']:
            period = record['期号']
            if period in seen_periods:
                continue
            
            seen_periods.add(period)
            record_count += 1
            
            # 获取位置详情
            play_categories = record.get('玩法分类', [])
            
            account_directions = []
            for idx, (account, direction, amount) in enumerate(zip(record['账户组'], record['方向组'], record['金额组'])):
                if '-' in direction:
                    clean_direction = direction.split('-', 1)[1]
                else:
                    clean_direction = direction
                
                # 显示位置详情
                if idx < len(play_categories):
                    position = play_categories[idx]
                    # 不再添加额外的位置分配信息
                    account_directions.append(f"{account}({position},{clean_direction}:¥{amount})")
                else:
                    account_directions.append(f"{account}({clean_direction}:¥{amount})")
            
            # 移除所有的位置分配和位置覆盖信息
            coverage_text = ""
            
            # 不再显示位置分配信息，因为已经在账户方向中包含了
            # coverage_text = ""
            
            if detect_type == 'PK10序列位置':
                st.write(f"{record_count}. 期号: {record['期号']} | 方向: {' ↔ '.join(account_directions)}")
            else:
                similarity_display = f"{record['相似度']:.2%}" if '相似度' in record else "100.00%"
                st.write(f"{record_count}. 期号: {record['期号']} | 方向: {' ↔ '.join(account_directions)} | 匹配度: {similarity_display}")
        
        if index < len(pattern):
            st.markdown("---")

    def display_summary_statistics(self, patterns):
        """显示总体统计"""
        if not patterns:
            return
            
        st.subheader("📊 总体统计")
        
        total_groups = len(patterns)
        total_accounts = sum(p['账户数量'] for p in patterns)
        total_wash_periods = sum(p['对刷期数'] for p in patterns)
        total_amount = sum(p['总投注金额'] for p in patterns)
        
        account_count_stats = defaultdict(int)
        for pattern in patterns:
            account_count_stats[pattern['账户数量']] += 1
        
        lottery_stats = defaultdict(int)
        for pattern in patterns:
            lottery_stats[pattern['彩种']] += 1
        
        activity_stats = defaultdict(int)
        for pattern in patterns:
            activity_stats[pattern['账户活跃度']] += 1
        
        opposite_type_stats = defaultdict(int)
        for pattern in patterns:
            for opposite_type, count in pattern['对立类型分布'].items():
                opposite_type_stats[opposite_type] += count
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("总对刷组数", total_groups)
        
        with col2:
            st.metric("涉及账户数", total_accounts)
        
        with col3:
            st.metric("总对刷期数", total_wash_periods)
        
        with col4:
            st.metric("总涉及金额", f"¥{total_amount:,.2f}")
        
        st.subheader("🎲 彩种类型统计")
        
        lottery_display_names = {
            'PK10': 'PK10/赛车',
            'K3': '快三',
            'LHC': '六合彩', 
            'SSC': '时时彩',
            '3D': '3D系列'
        }
        
        lottery_cols = st.columns(min(5, len(lottery_stats)))
        
        for i, (lottery, count) in enumerate(lottery_stats.items()):
            if i < len(lottery_cols):
                with lottery_cols[i]:
                    display_name = lottery_display_names.get(lottery, lottery)
                    st.metric(
                        label=display_name,
                        value=f"{count}组"
                    )
        
        col_left, col_right = st.columns(2)
        
        with col_left:
            st.subheader("👥 账户组合分布")
            
            for account_count, group_count in sorted(account_count_stats.items()):
                account_type_periods = sum(p['对刷期数'] for p in patterns if p['账户数量'] == account_count)
                st.write(f"- **{account_count}组**: {group_count}组 ({account_type_periods}期)")
        
        with col_right:
            st.subheader("📈 活跃度分布")
            
            activity_display_names = {
                'low': '低活跃度',
                'medium': '中活跃度',
                'high': '高活跃度',
                'very_high': '极高活跃度'
            }
            
            for activity, count in activity_stats.items():
                display_name = activity_display_names.get(activity, activity)
                activity_periods = sum(p['对刷期数'] for p in patterns if p['账户活跃度'] == activity)
                st.write(f"- **{display_name}**: {count}组 ({activity_periods}期)")
        
        st.subheader("📈 关键指标")
        
        avg_group_amount = total_amount / total_groups if total_groups > 0 else 0
        
        metric_col1, metric_col2, metric_col3 = st.columns(3)
        
        with metric_col1:
            st.metric("平均每组金额", f"¥{avg_group_amount:,.2f}")
        
        with metric_col2:
            business_total = total_amount
            st.metric("业务类型总额", f"¥{business_total:,.2f}")
        
        with metric_col3:
            st.metric("参与总账户数", total_accounts)
        
        st.subheader("🎯 主要对立类型")
        
        top_opposites = sorted(opposite_type_stats.items(), key=lambda x: x[1], reverse=True)[:3]
        
        for opposite_type, count in top_opposites:
            if ' vs ' in opposite_type:
                display_type = opposite_type.replace(' vs ', '-')
            else:
                display_type = opposite_type
            st.write(f"- **{display_type}**: {count}期")

    def export_detection_results(self, patterns, export_format='excel'):
        """导出检测结果"""
        if not patterns:
            st.warning("❌ 没有检测结果可供导出")
            return None
        
        try:
            main_data = []
            detailed_data = []
            
            for i, pattern in enumerate(patterns, 1):
                main_record = {
                    '组ID': f"组{i}",
                    '账户组': ' ↔ '.join(pattern['账户组']),
                    '彩种': pattern['彩种'],
                    '彩种类型': pattern['彩种类型'],
                    '账户数量': pattern['账户数量'],
                    '主要对立类型': pattern['主要对立类型'],
                    '对刷期数': pattern['对刷期数'],
                    '要求最小对刷期数': pattern['要求最小对刷期数'],
                    '总投注金额': pattern['总投注金额'],
                    '平均相似度': pattern['平均相似度'],
                    '账户活跃度': pattern['账户活跃度'],
                    '账户统计信息': '; '.join(pattern['账户统计信息'])
                }
                main_data.append(main_record)
                
                for j, record in enumerate(pattern['详细记录'], 1):
                    detailed_record = {
                        '组ID': f"组{i}",
                        '账户组': ' ↔ '.join(pattern['账户组']),
                        '期号': record['期号'],
                        '彩种': record['彩种'],
                        '彩种类型': record['彩种类型'],
                        '方向组': ' ↔ '.join([f"{acc}({dir})" for acc, dir in zip(record['账户组'], record['方向组'])]),
                        '金额组': ' ↔ '.join([f"¥{amt}" for amt in record['金额组']]),
                        '总金额': record['总金额'],
                        '相似度': record['相似度'],
                        '账户数量': record['账户数量'],
                        '模式': record['模式'],
                        '对立类型': record['对立类型']
                    }
                    detailed_data.append(detailed_record)
            
            df_main = pd.DataFrame(main_data)
            df_detailed = pd.DataFrame(detailed_data)
            
            numeric_columns = ['总投注金额', '平均相似度', '总金额', '相似度']
            for col in numeric_columns:
                if col in df_main.columns:
                    df_main[col] = df_main[col].apply(lambda x: f"¥{x:,.2f}" if '金额' in col else f"{x:.2%}")
                if col in df_detailed.columns:
                    df_detailed[col] = df_detailed[col].apply(lambda x: f"¥{x:,.2f}" if '金额' in col else f"{x:.2%}")
            
            if export_format == 'excel':
                return self._export_to_excel(df_main, df_detailed)
            else:
                return self._export_to_csv(df_main, df_detailed)
                
        except Exception as e:
            logger.error(f"导出失败: {str(e)}")
            st.error(f"导出失败: {str(e)}")
            return None

    def _export_to_excel(self, df_main, df_detailed):
        """导出到Excel格式"""
        try:
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_main.to_excel(writer, sheet_name='对刷组汇总', index=False)
                df_detailed.to_excel(writer, sheet_name='详细记录', index=False)
                
                workbook = writer.book
                main_sheet = workbook['对刷组汇总']
                detailed_sheet = workbook['详细记录']
                
                for sheet in [main_sheet, detailed_sheet]:
                    for column in sheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        sheet.column_dimensions[column_letter].width = adjusted_width
                
                main_sheet.insert_rows(0, 3)
                main_sheet['A1'] = "对刷检测结果报告"
                main_sheet['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                main_sheet['A3'] = f"总对刷组数: {len(df_main)}"
                
                main_sheet.merge_cells('A1:L1')
                main_sheet.merge_cells('A2:L2')
                main_sheet.merge_cells('A3:L3')
                
                for cell in ['A1', 'A2', 'A3']:
                    main_sheet[cell].font = Font(bold=True, size=12)
                    main_sheet[cell].alignment = Alignment(horizontal='center')
            
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Excel导出失败: {str(e)}")
            raise e

    def _export_to_csv(self, df_main, df_detailed):
        """导出到CSV格式"""
        try:
            zip_buffer = io.BytesIO()
            
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                main_csv = df_main.to_csv(index=False, encoding='utf-8-sig')
                zip_file.writestr('对刷组汇总.csv', main_csv)
                
                detailed_csv = df_detailed.to_csv(index=False, encoding='utf-8-sig')
                zip_file.writestr('详细记录.csv', detailed_csv)
                
                readme_content = f"""对刷检测结果导出文件
生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
总对刷组数: {len(df_main)}

文件说明:
1. 对刷组汇总.csv - 包含所有对刷组的汇总信息
2. 详细记录.csv - 包含每个对刷组的详细期号记录

检测参数:
- 最小投注金额: {self.config.min_amount}元
- 基础匹配度阈值: {self.config.amount_similarity_threshold:.0%}
- 最大检测账户数: {self.config.max_accounts_in_group}
"""
                zip_file.writestr('说明.txt', readme_content)
            
            zip_buffer.seek(0)
            return zip_buffer
            
        except Exception as e:
            logger.error(f"CSV导出失败: {str(e)}")
            raise e

    def display_export_buttons(self, patterns):
        """显示导出按钮"""
        if not patterns:
            return
        
        st.markdown("---")
        st.subheader("📤 导出检测结果")
        
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("📊 导出Excel报告", use_container_width=True):
                with st.spinner("正在生成Excel报告..."):
                    excel_data = self.export_detection_results(patterns, 'excel')
                    if excel_data:
                        st.download_button(
                            label="⬇️ 下载Excel文件",
                            data=excel_data,
                            file_name=f"对刷检测报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
        
        with col2:
            if st.button("📄 导出CSV文件", use_container_width=True):
                with st.spinner("正在生成CSV文件..."):
                    csv_data = self.export_detection_results(patterns, 'csv')
                    if csv_data:
                        st.download_button(
                            label="⬇️ 下载CSV压缩包",
                            data=csv_data,
                            file_name=f"对刷检测报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                            mime="application/zip",
                            use_container_width=True
                        )
        
        st.info(f"📊 导出内容: {len(patterns)}个对刷组, 共{sum(len(p['详细记录']) for p in patterns)}条详细记录")

# ==================== 主函数 ====================
def main():
    """主函数"""
    st.title("🎯 🎈智能对刷检测系统🎈")
    st.markdown("---")
    
    with st.sidebar:
        st.header("📁 数据上传")
        uploaded_file = st.file_uploader(
            "上传投注数据文件", 
            type=['xlsx', 'xls', 'csv'],
            help="请上传包含彩票投注数据的Excel或CSV文件"
        )
        
        st.header("⚙️ 检测参数设置")
        
        min_amount = st.slider(
            "最小投注金额阈值", 
            min_value=1, 
            max_value=50, 
            value=5,
            help="投注金额低于此值的记录将不参与检测"
        )
        
        max_accounts = st.slider(
            "最大检测账户数", 
            2, 8, 4, 
            help="检测的最大账户组合数量"
        )
        
        period_diff_threshold = st.slider(
            "账户期数最大差异阈值", 
            min_value=0, 
            max_value=500,
            value=100,
            help="账户总投注期数最大允许差异，超过此值不进行组合检测"
        )
        
        st.subheader("💰 金额平衡设置")
        
        enable_balance_filter = st.checkbox("启用金额平衡过滤", value=True,
                                          help="确保对刷组内账户金额差距不超过设定倍数")
        
        max_ratio = 10
        if enable_balance_filter:
            max_ratio = st.slider("最大金额差距倍数", 
                                 min_value=1, 
                                 max_value=20, 
                                 value=5, 
                                 step=1,
                                 help="组内最大金额与最小金额的允许倍数（例如：10表示10倍差距）")
        
        st.subheader("🎯 多账户匹配度配置")
        
        st.markdown("**2个账户:**")
        similarity_2_accounts = st.slider(
            "2个账户匹配度阈值", 
            min_value=0.3, max_value=1.0, value=0.7, step=0.01,
            help="2个账户对刷的金额匹配度阈值"
        )
        
        st.markdown("**3个账户:**")
        similarity_3_accounts = st.slider(
            "3个账户匹配度阈值", 
            min_value=0.3, max_value=1.0, value=0.8, step=0.01,
            help="3个账户对刷的金额匹配度阈值"
        )
        
        st.markdown("**4个账户:**")
        similarity_4_accounts = st.slider(
            "4个账户匹配度阈值", 
            min_value=0.3, max_value=1.0, value=0.85, step=0.01,
            help="4个账户对刷的金额匹配度阈值"
        )
        
        st.markdown("**5个账户:**")
        similarity_5_accounts = st.slider(
            "5个账户匹配度阈值", 
            min_value=0.3, max_value=1.0, value=0.9, step=0.01,
            help="5个账户对刷的金额匹配度阈值"
        )
        
        st.subheader("🛠️ 连续对刷阈值配置")
        
        st.markdown("**低活跃度(1-10期):**")
        min_periods_low = st.slider(
            "低活跃度最小连续对刷期数", 
            min_value=1, max_value=10, value=3,
            help="总投注期数1-10期的账户，要求的最小连续对刷期数"
        )
        
        st.markdown("**中活跃度(11-50期):**")
        min_periods_medium = st.slider(
            "中活跃度最小连续对刷期数", 
            min_value=3, max_value=15, value=5,
            help="总投注期数11-50期的账户，要求的最小连续对刷期数"
        )
        
        st.markdown("**高活跃度(51-100期):**")
        min_periods_high = st.slider(
            "高活跃度最小连续对刷期数", 
            min_value=5, max_value=20, value=8,
            help="总投注期数51-100期的账户，要求的最小连续对刷期数"
        )
        
        st.markdown("**极高活跃度(100期以上):**")
        min_periods_very_high = st.slider(
            "极高活跃度最小连续对刷期数", 
            min_value=8, max_value=30, value=11,
            help="总投注期数100期以上的账户，要求的最小连续对刷期数"
        )
    
    if uploaded_file is not None:
        try:
            config = Config()
            config.min_amount = min_amount
            config.max_accounts_in_group = max_accounts
            config.account_period_diff_threshold = period_diff_threshold
            
            config.amount_similarity_threshold = similarity_2_accounts
            
            config.amount_threshold = {
                'max_amount_ratio': max_ratio,
                'enable_threshold_filter': enable_balance_filter
            }
            
            config.account_count_similarity_thresholds = {
                2: similarity_2_accounts,
                3: similarity_3_accounts,
                4: similarity_4_accounts,
                5: similarity_5_accounts
            }
            
            config.period_thresholds.update({
                'min_periods_low': min_periods_low,
                'min_periods_medium': min_periods_medium,
                'min_periods_high': min_periods_high,
                'min_periods_very_high': min_periods_very_high
            })
            
            detector = WashTradeDetector(config)
            
            st.success(f"✅ 已上传文件: {uploaded_file.name}")
            
            with st.spinner("🔄 正在解析数据..."):
                df_enhanced, filename = detector.upload_and_process(uploaded_file)
                
                if df_enhanced is not None and len(df_enhanced) > 0:
                    with st.spinner("🔍 正在检测对刷交易..."):
                        patterns = detector.detect_all_wash_trades()
                    
                    if patterns:
                        detector.display_detailed_results(patterns)
                        detector.display_export_buttons(patterns)
                    else:
                        st.warning("⚠️ 未发现符合阈值条件的对刷行为")
                else:
                    st.error("❌ 数据解析失败，请检查文件格式和内容")
            
        except Exception as e:
            st.error(f"❌ 程序执行失败: {str(e)}")
    else:
        st.info("👈 请在左侧边栏上传数据文件开始分析")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.subheader("🔍 智能检测")
            st.markdown("""
            - 多账户对刷模式识别
            - 智能金额匹配分析
            - 活跃度自适应阈值
            - 实时进度监控
            """)
        
        with col2:
            st.subheader("📊 专业分析")
            st.markdown("""
            - 完整彩种支持
            - 玩法分类标准化
            - 数据质量验证
            - 详细统计报告
            """)
        
        with col3:
            st.subheader("🚀 高效处理")
            st.markdown("""
            - 大数据量优化
            - 并行检测算法
            - 一键导出结果
            - 实时性能监控
            """)
    
    with st.expander("📖 系统使用说明", expanded=False):
        st.markdown("""
        ### 系统功能说明

        **🎯 检测逻辑：**
        - **金额过滤**：投注金额低于设定阈值（默认10元）的记录不参与检测
        - **总投注期数**：账户在特定彩种中的所有期号投注次数
        - **对刷期数**：账户组实际发生对刷行为的期数
        - 根据**总投注期数**判定账户活跃度，设置不同的**对刷期数**阈值

        **📊 活跃度判定：**
        - **1-10期**：要求≥3期连续对刷
        - **11-50期**：要求≥5期连续对刷  
        - **51-100期**：要求≥8期连续对刷
        - **100期以上**：要求≥11期连续对刷

        **🎯 多账户匹配度要求：**
        - **2个账户**：80%匹配度
        - **3个账户**：85%匹配度  
        - **4个账户**：90%匹配度
        - **5个账户**：95%匹配度

        **🔄 账户期数差异检查：**
        - 避免期数差异过大的账户组合
        - 默认阈值：101期
        - 可自定义调整阈值

        **⚡ 自动检测：**
        - 数据上传后自动开始处理和分析
        - 无需手动点击开始检测按钮

        **🎲 新增六合彩检测：**
        - **天肖 vs 地肖**：天肖与地肖的对立检测
        - **家肖 vs 野肖**：家禽肖与野兽肖的对立检测  
        - **尾大 vs 尾小**：尾数大小的对立检测
        - **特大 vs 特小**：特码大小的对立检测
        - **特单 vs 特双**：特码单双的对立检测
        """)

if __name__ == "__main__":
    main()
